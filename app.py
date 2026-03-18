"""
╔══════════════════════════════════════════════════════════════════╗
║              FinShield OSINT v3  —  Architecture Guide           ║
╠══════════════════════════════════════════════════════════════════╣
║  SECTIONS (search "# §" to jump between them)                    ║
║                                                                  ║
║  § 0  IMPORTS & PAGE CONFIG                                      ║
║  § 1  CONFIGURATION  ← edit thresholds, query lists here        ║
║  § 2  DATABASE LAYER  (SQLite — banks, IBAN, reports, watchlist) ║
║  § 3  IBAN VALIDATION                                            ║
║  § 4  OSINT ENGINE    ← search queries + scoring logic here      ║
║  § 5  PDF REPORT      ← ReportLab layout here                   ║
║  § 6  EXCEL HISTORY   ← openpyxl workbook logic here            ║
║  § 7  STREAMLIT SIDEBAR                                          ║
║  § 8  TAB 1 — IBAN VERIFICATION                                  ║
║  § 9  TAB 2 — OSINT ANALYSIS                                     ║
║  § 10 TAB 3 — BANK SEARCH                                        ║
║  § 11 TAB 4 — DATABASE MANAGEMENT                                ║
║  § 12 TAB 5 — HISTORY & WATCHLIST                                ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ─────────────────────────────────────────────────────────────────
# § 0  IMPORTS & PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
import streamlit as st
import requests, json, re, time, os, sqlite3, csv, io, textwrap
from datetime import datetime
from io import BytesIO
from urllib.parse import quote_plus, urlparse

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable, PageBreak)
from reportlab.lib.units import mm

st.set_page_config(
    page_title="FinShield OSINT v3",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS (unchanged from v2) ───────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');
:root{--bg:#0a0d14;--surface:#111520;--border:#1e2535;--accent:#00d4ff;
  --accent2:#ff6b35;--green:#00ff88;--red:#ff3366;--yellow:#ffcc00;
  --text:#c8d6e5;--muted:#5a6a7a;}
html,body,[data-testid="stApp"]{background:var(--bg)!important;color:var(--text);font-family:'IBM Plex Sans',sans-serif;}
[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border);}
[data-testid="stSidebar"] *{color:var(--text)!important;}
h1,h2,h3{font-family:'IBM Plex Mono',monospace!important;}
h1{color:var(--accent)!important;}
h2{color:var(--text)!important;border-bottom:1px solid var(--border);padding-bottom:8px;}
h3{color:var(--accent)!important;font-size:0.95rem!important;}
input,textarea,[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{
  background:var(--surface)!important;color:var(--text)!important;
  border:1px solid var(--border)!important;border-radius:4px!important;
  font-family:'IBM Plex Mono',monospace!important;}
.stButton>button{background:transparent!important;color:var(--accent)!important;
  border:1px solid var(--accent)!important;border-radius:3px!important;
  font-family:'IBM Plex Mono',monospace!important;font-size:0.85rem!important;
  letter-spacing:1px!important;transition:all 0.15s!important;}
.stButton>button:hover{background:var(--accent)!important;color:var(--bg)!important;}
.metric-card{background:var(--surface);border:1px solid var(--border);border-radius:6px;padding:16px 20px;margin:8px 0;}
.metric-card .label{font-size:0.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:2px;}
.metric-card .value{font-family:'IBM Plex Mono',monospace;font-size:1.4rem;color:var(--accent);margin-top:4px;}
.badge-low{background:rgba(0,255,136,0.1);color:var(--green);border:1px solid rgba(0,255,136,0.3);padding:2px 10px;border-radius:12px;font-size:0.75rem;}
.badge-medium{background:rgba(255,204,0,0.1);color:var(--yellow);border:1px solid rgba(255,204,0,0.3);padding:2px 10px;border-radius:12px;font-size:0.75rem;}
.badge-high{background:rgba(255,51,102,0.1);color:var(--red);border:1px solid rgba(255,51,102,0.3);padding:2px 10px;border-radius:12px;font-size:0.75rem;}
.info-box{background:rgba(0,212,255,0.05);border-left:3px solid var(--accent);padding:12px 16px;margin:8px 0;border-radius:0 4px 4px 0;font-size:0.88rem;}
.warn-box{background:rgba(255,204,0,0.05);border-left:3px solid var(--yellow);padding:12px 16px;margin:8px 0;border-radius:0 4px 4px 0;font-size:0.88rem;}
.danger-box{background:rgba(255,51,102,0.07);border-left:3px solid var(--red);padding:12px 16px;margin:8px 0;border-radius:0 4px 4px 0;font-size:0.88rem;}
.ok-box{background:rgba(0,255,136,0.05);border-left:3px solid var(--green);padding:12px 16px;margin:8px 0;border-radius:0 4px 4px 0;font-size:0.88rem;}
.result-row{background:var(--surface);border:1px solid var(--border);border-radius:4px;padding:12px 16px;margin:6px 0;font-size:0.85rem;}
[data-testid="stTabs"] button{background:transparent!important;color:var(--muted)!important;
  border-radius:0!important;border-bottom:2px solid transparent!important;
  font-family:'IBM Plex Mono',monospace!important;font-size:0.8rem!important;letter-spacing:1px!important;}
[data-testid="stTabs"] button[aria-selected="true"]{color:var(--accent)!important;border-bottom-color:var(--accent)!important;background:rgba(0,212,255,0.05)!important;}
[data-testid="stExpander"]{background:var(--surface)!important;border:1px solid var(--border)!important;border-radius:4px!important;}
hr{border-color:var(--border)!important;}
.stProgress>div>div{background:var(--accent)!important;}
.header-strip{display:flex;align-items:center;gap:16px;padding:20px 0 12px;border-bottom:1px solid var(--border);margin-bottom:24px;}
.app-title{font-family:'IBM Plex Mono',monospace;font-size:1.8rem;color:var(--accent);}
.app-sub{font-size:0.8rem;color:var(--muted);letter-spacing:2px;text-transform:uppercase;}
.section-title{font-family:'IBM Plex Mono',monospace;font-size:1rem;color:var(--accent);
  text-transform:uppercase;letter-spacing:2px;margin:20px 0 10px;padding-bottom:6px;border-bottom:1px solid var(--border);}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# § 1  CONFIGURATION
# ─────────────────────────────────────────────────────────────────
# ── Score weights per alert category ─────────────────────────────
SCORE_WEIGHTS = {
    "sanctions": 3.0,   # Heaviest — confirmed lists
    "fraud":     2.5,   # Financial crime
    "judicial":  2.0,   # Court proceedings
    "reputation":1.5,   # Press, reviews
    "pep":       1.0,   # PEP exposure
}

# ── Source credibility multipliers ───────────────────────────────
SOURCE_CREDIBILITY = {
    # Official/regulatory
    "opensanctions.org":2.0,"legifrance.gouv.fr":2.0,"justice.fr":2.0,
    "bodacc.fr":1.8,"amf-france.org":2.0,"acpr.banque-france.fr":2.0,
    "interpol.int":2.0,"europol.europa.eu":2.0,"tracfin.gouv.fr":2.0,
    "courdecassation.fr":1.9,"sec.gov":2.0,"ofac.treas.gov":2.0,
    "pacer.gov":1.8,"companieshouse.gov.uk":1.7,"opencorporates.com":1.5,
    "pappers.fr":1.6,"societe.com":1.4,"infogreffe.fr":1.8,
    # Press — major
    "lemonde.fr":1.5,"lefigaro.fr":1.5,"liberation.fr":1.3,
    "bfmtv.com":1.3,"franceinfo.fr":1.4,"latribune.fr":1.4,
    "lesechos.fr":1.5,"capital.fr":1.3,"leparisien.fr":1.3,
    "reuters.com":1.6,"bloomberg.com":1.6,"ft.com":1.6,
    "theguardian.com":1.5,"nytimes.com":1.5,"bbc.com":1.4,
    # Review/scam reporting
    "trustpilot.com":1.2,"signal-arnaques.com":1.5,
    "cybermalveillance.gouv.fr":1.8,"avis-verifies.com":1.1,
    # Social (lower weight)
    "twitter.com":0.9,"linkedin.com":1.0,"reddit.com":0.9,
    "facebook.com":0.8,"instagram.com":0.8,
}

# ── Score → risk level thresholds ────────────────────────────────
RISK_THRESHOLDS = {"CRITIQUE": 70, "ELEVE": 40, "MODERE": 10, "FAIBLE": 0}

# ── Gravity weights used in query-result scoring ──────────────────
GRAVITY_WEIGHTS = {"eleve": 5.0, "moyen": 3.0, "faible": 1.0}

# ── Excel history file path ───────────────────────────────────────
EXCEL_HISTORY_PATH = "finshield_history.xlsx"

# ── SQLite database path ──────────────────────────────────────────
DB_PATH = "finshield.db"


# ─────────────────────────────────────────────────────────────────
# § 2  DATABASE LAYER
# ─────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS banks (
        code TEXT PRIMARY KEY, name TEXT NOT NULL,
        address TEXT DEFAULT '', city TEXT DEFAULT '',
        postal_code TEXT DEFAULT '', country TEXT DEFAULT 'FR',
        bic TEXT DEFAULT '', type TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS iban_countries (
        code TEXT PRIMARY KEY, name TEXT NOT NULL, length INTEGER NOT NULL,
        structure TEXT DEFAULT '', example TEXT DEFAULT '',
        bban_format TEXT DEFAULT '', notes TEXT DEFAULT '')""")
    c.execute("""CREATE TABLE IF NOT EXISTS osint_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entity TEXT NOT NULL, entity_type TEXT DEFAULT '',
        iban TEXT DEFAULT '', score INTEGER DEFAULT 0,
        niveau TEXT DEFAULT '', recommandation TEXT DEFAULT '',
        resume TEXT DEFAULT '', full_json TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS watchlist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entity TEXT NOT NULL, entity_type TEXT DEFAULT '',
        reason TEXT DEFAULT '', risk_level TEXT DEFAULT '',
        added_by TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit(); conn.close()

def seed_banks():
    """
    Charge les banques depuis swift_banks.csv si présent dans le même répertoire
    que l'app, sinon utilise la liste FR intégrée en fallback.
    Format CSV : code,name,address,city,postal_code,country,bic,type
    """
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM banks")
    if c.fetchone()[0] > 0: conn.close(); return

    banks = []

    # ── 1. Essai depuis swift_banks.csv (même dossier que app.py) ──────────
    import os, csv as csv_module
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "swift_banks.csv")
    if os.path.exists(csv_path):
        try:
            with open(csv_path, newline='', encoding='utf-8') as f:
                reader = csv_module.DictReader(f)
                for row in reader:
                    banks.append((
                        row.get('code',''),
                        row.get('name',''),
                        row.get('address',''),
                        row.get('city',''),
                        row.get('postal_code',''),
                        row.get('country',''),
                        row.get('bic',''),
                        row.get('type','Etablissement de crédit'),
                        ''  # notes
                    ))
        except Exception:
            banks = []  # fallback si CSV corrompu

    # ── 2. Fallback : banques françaises hardcodées ─────────────────────────
    if not banks:
        banks = [
            ('18989','Aareal bank AG','29 B RUE D ASTORG','PARIS 08','75008','FR','','Etablissement de crédit',''),
            ('30004','BNP Paribas','16 BOULEVARD DES ITALIENS','PARIS','75009','FR','BNPAFRPP','Etablissement de crédit',''),
            ('30001','BANQUE DE FRANCE','1 RUE LA VRILLIERE','PARIS 01','75001','FR','BDFEFRPP','Banque centrale',''),
            ('30003','Societe generale','189 RUE D AUBERVILLIERS','PARIS CEDEX 18','75886','FR','SOGEFRPP','Etablissement de crédit',''),
            ('30002','CREDIT LYONNAIS LCL','18 rue de la Republique','LYON','69002','FR','CRLYFRPP','Etablissement de crédit',''),
            ('30006','Credit Agricole S.A.','12 PLACE DES ETATS UNIS','MONTROUGE','92120','FR','AGRIFRPP','Etablissement de crédit',''),
            ('40618','Boursorama','18 QUAI DU POINT DU JOUR','BOULOGNE BILLANCOURT','92100','FR','BOUSFRPP','Etablissement de crédit',''),
            ('20041','La Banque Postale','115 RUE DE SEVRES','PARIS CEDEX 06','75275','FR','PSSTFRPP','Etablissement de crédit',''),
            ('16188','BPCE','50 AVENUE PIERRE MENDES FRANCE','PARIS 13','75013','FR','BPCEFRPP','Etablissement de crédit',''),
            ('10107','BRED Banque populaire','18 QUAI DE LA RAPEE','PARIS 12','75012','FR','BREDFRPP','Etablissement de crédit',''),
            ('30007','Natixis','30 AVENUE PIERRE MENDES FRANCE','PARIS 13','75013','FR','NATXFRPP','Etablissement de crédit',''),
            ('30066','Credit industriel et commercial CIC','6 AVENUE DE PROVENCE','PARIS 09','75009','FR','CMCIFRPP','Etablissement de crédit',''),
            ('30076','Credit du Nord','28 PLACE RIHOUR','LILLE CEDEX','59023','FR','NORDFRPP','Etablissement de crédit',''),
            ('41189','Banco Bilbao Vizcaya Argentaria (BBVA)','29 AVENUE DE L OPERA','PARIS 01','75001','FR','BBVAFRPP','Etablissement de crédit',''),
            ('44729','Banco Santander SA','40 RUE DE COURCELLES','PARIS 08','75008','FR','BSCHFRPP','Etablissement de crédit',''),
            ('18769','Bank of China limited','23 AVENUE DE LA GRANDE ARMEE','PARIS 16','75116','FR','BKCHFRPP','Etablissement de crédit',''),
            ('11833','ICBC Europe SA','73 BOULEVARD HAUSSMANN','PARIS 08','75008','FR','ICBKFRPP','Etablissement de crédit',''),
            ('30438','ING Bank NV','40 AVENUE DES TERROIRS DE FRANCE','PARIS 12','75012','FR','INGBFRPP','Etablissement de crédit',''),
            ('30628','JPMorgan Chase bank','14 PLACE VENDOME','PARIS 01','75001','FR','CHASFRPP','Etablissement de crédit',''),
            ('25533','Goldman Sachs Bank Europe SE','5 Avenue Kleber','PARIS','75116','FR','GOLDFRPP','Etablissement de crédit',''),
            ('30056','HSBC Continental Europe','38 AVENUE KLEBER','PARIS','75116','FR','CCFRFRPP','Etablissement de crédit',''),
            ('30758','UBS France S.A.','69 BOULEVARD HAUSSMANN','PARIS 08','75008','FR','UBSWFRPP','Etablissement de crédit',''),
            ('42559','Credit cooperatif','12 BOULEVARD PESARO','NANTERRE CEDEX','92024','FR','CCOPFRPP','Etablissement de crédit',''),
            ('30788','Banque Neuflize OBC','3 AVENUE HOCHE','PARIS 08','75008','FR','NEIOFR22','Etablissement de crédit',''),
            ('12548','Axa banque','203-205 RUE CARNOT','FONTENAY-SOUS-BOIS CEDEX','94138','FR','AXABFRPP','Etablissement de crédit',''),
            ('11188','RCI Banque','15 RUE D UZES','PARIS','75002','FR','RCIEFR22','Etablissement de crédit',''),
            ('30748','Lazard Freres Banque','121 BOULEVARD HAUSSMANN','PARIS 08','75008','FR','LAZAFRPP','Etablissement de crédit',''),
        ]

    c.executemany("""INSERT OR IGNORE INTO banks
        (code,name,address,city,postal_code,country,bic,type,notes) VALUES (?,?,?,?,?,?,?,?,?)""", banks)
    conn.commit(); conn.close()

def seed_iban_countries():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM iban_countries")
    if c.fetchone()[0] > 0: conn.close(); return
    countries = [
        ("AL","Albanie",28,"ALkk bbbs sssx cccc cccc cccc cccc","AL47212110090000000235698741","8n,16c",""),
        ("AD","Andorre",24,"ADkk bbbb ssss cccc cccc cccc","AD1200012030200359100100","4n,4n,12c",""),
        ("AT","Autriche",20,"ATkk bbbb bccc cccc cccc","AT611904300234573201","5n,11n",""),
        ("BE","Belgique",16,"BEkk bbbc cccc ccxx","BE62510007547061","3n,7n,2n",""),
        ("BG","Bulgarie",22,"BGkk bbbb ssss ttcc cccc cc","BG80BNBG96611020345678","4a,4n,2n,8c",""),
        ("CH","Suisse",21,"CHkk bbbb bccc cccc cccc c","CH9300762011623852977","5n,12c",""),
        ("CY","Chypre",28,"CYkk bbbs ssss cccc cccc cccc cccc","CY17002001280000001200527600","3n,5n,16c",""),
        ("CZ","Rép. Tchèque",24,"CZkk bbbb ssss sscc cccc cccc","CZ6508000000192000145399","4n,6n,10n",""),
        ("DE","Allemagne",22,"DEkk bbbb bbbb cccc cccc cc","DE89370400440532013000","8n,10n",""),
        ("DK","Danemark",18,"DKkk bbbb cccc cccc cc","DK5000400440116243","4n,9n,1n",""),
        ("EE","Estonie",20,"EEkk bbss cccc cccc cccx","EE382200221020145399","2n,2n,11n,1n",""),
        ("ES","Espagne",24,"ESkk bbbb gggg xxcc cccc cccc","ES9121000418450200051332","4n,4n,1n,1n,10n",""),
        ("FI","Finlande",18,"FIkk bbbb bbcc cccc cx","FI2112345600000785","6n,7n,1n",""),
        ("FR","France",27,"FRkk bbbb bggg ggcc cccc cccc cxx","FR7630006000011234567890189","5n,5n,11c,2n",""),
        ("GB","Grande-Bretagne",22,"GBkk bbbb ssss sscc cccc cc","GB29NWBK60161331926819","4a,6n,8n",""),
        ("GR","Grèce",27,"GRkk bbbs sssc cccc cccc cccc ccc","GR1601101250000000012300695","3n,4n,16n",""),
        ("HR","Croatie",21,"HRkk bbbb bbbc cccc cccc c","HR1210010051863000160","7n,10n",""),
        ("HU","Hongrie",28,"HUkk bbbs sssk cccc cccc cccc cccx","HU42117730161111101800000000","3n,4n,1n,15n,1n",""),
        ("IE","Irlande",22,"IEkk aaaa bbbb bbcc cccc cc","IE29AIBK93115212345678","4a,6n,8n",""),
        ("IS","Islande",26,"ISkk bbbb sscc cccc iiii iiii ii","IS140159260076545510730339","4n,2n,6n,10n",""),
        ("IT","Italie",27,"ITkk xbbb bbss sssc cccc cccc ccc","IT60X0542811101000000123456","1a,5n,5n,12c",""),
        ("LI","Liechtenstein",21,"LIkk bbbb bccc cccc cccc c","LI21088100002324013AA","5n,12c",""),
        ("LT","Lituanie",20,"LTkk bbbb bccc cccc cccc","LT121000011101001000","5n,11n",""),
        ("LU","Luxembourg",20,"LUkk bbbc cccc cccc cccc","LU280019400644750000","3n,13c",""),
        ("LV","Lettonie",21,"LVkk bbbb cccc cccc cccc c","LV80BANK0000435195001","4a,13c",""),
        ("MC","Monaco",27,"MCkk bbbb bggg ggcc cccc cccc cxx","MC5811222000010123456789030","5n,5n,11c,2n",""),
        ("MT","Malte",31,"MTkk bbbb ssss sccc cccc cccc cccc ccc","MT84MALT011000012345MTLCAST001S","4a,5n,18c",""),
        ("NL","Pays-Bas",18,"NLkk bbbb cccc cccc cc","NL39RABO0300065264","4a,10n",""),
        ("NO","Norvège",15,"NOkk bbbb cccc ccx","NO9386011117947","4n,6n,1n",""),
        ("PL","Pologne",28,"PLkk bbbs sssx cccc cccc cccc cccc","PL27114020040000300201355387","8n,16n",""),
        ("PT","Portugal",25,"PTkk bbbb ssss cccc cccc cccx x","PT50000201231234567890154","4n,4n,11n,2n",""),
        ("RO","Roumanie",24,"ROkk bbbb cccc cccc cccc cccc","RO49AAAA1B31007593840000","4a,16c",""),
        ("SE","Suède",24,"SEkk bbbc cccc cccc cccc cccc","SE3550000000054910000003","3n,16n,1n",""),
        ("SI","Slovénie",19,"SIkk bbss sccc cccc cxx","SI56191000000123438","5n,8n,2n",""),
        ("SK","Slovaquie",24,"SKkk bbbb ssss sscc cccc cccc","SK3112000000198742637341","4n,6n,10n",""),
        ("SM","Saint-Marin",27,"SMkk xbbb bbss sssc cccc cccc ccc","SM86U0322509800000000270100","1a,5n,5n,12c",""),
        ("TR","Turquie",26,"TRkk bbbb bxcc cccc cccc cccc cc","TR330006100519786457841326","5n,1c,16c",""),
        ("MA","Maroc",28,"MAkk bbbb bsss sscc cccc cccc cccc","","3n,3n,2n,16n",""),
        ("TN","Tunisie",24,"TNkk bbss sccc cccc cccc cccc","","2n,3n,13n,2n",""),
        ("DZ","Algérie",26,"DZkk bbbb bsss sscc cccc cccc cc","","5n,5n,10n,2n",""),
        ("AE","Émirats Arabes Unis",23,"AEkk bbbc cccc cccc cccc ccc","","3n,16n",""),
        ("SA","Arabie Saoudite",24,"SAkk bbcc cccc cccc cccc cccc","","2n,18c",""),
    ]
    c.executemany("""INSERT OR IGNORE INTO iban_countries
        (code,name,length,structure,example,bban_format,notes) VALUES (?,?,?,?,?,?,?)""", countries)
    conn.commit(); conn.close()

# CRUD helpers
def db_get_banks(search=""):
    conn = get_db()
    if search:
        q = f"%{search.lower()}%"
        rows = conn.execute(
            "SELECT * FROM banks WHERE lower(name) LIKE ? OR code LIKE ? OR bic LIKE ? ORDER BY name",
            (q,q,q)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM banks ORDER BY name").fetchall()
    conn.close(); return [dict(r) for r in rows]

def db_get_bank_by_code(code):
    conn = get_db()
    row = conn.execute("SELECT * FROM banks WHERE code=? OR code=?",
                       (code, code.lstrip("0"))).fetchone()
    if not row:
        row = conn.execute("SELECT * FROM banks WHERE code LIKE ? LIMIT 1",
                           (code[:4]+"%",)).fetchone()
    conn.close(); return dict(row) if row else None

def db_upsert_bank(code, name, address, city, postal_code, country, bic, btype, notes):
    conn = get_db()
    conn.execute("""INSERT INTO banks (code,name,address,city,postal_code,country,bic,type,notes,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
        ON CONFLICT(code) DO UPDATE SET name=excluded.name,address=excluded.address,
        city=excluded.city,postal_code=excluded.postal_code,country=excluded.country,
        bic=excluded.bic,type=excluded.type,notes=excluded.notes,updated_at=CURRENT_TIMESTAMP""",
        (code,name,address,city,postal_code,country,bic,btype,notes))
    conn.commit(); conn.close()

def db_delete_bank(code):
    conn = get_db()
    conn.execute("DELETE FROM banks WHERE code=?", (code,))
    conn.commit(); conn.close()

def db_get_iban_country(code):
    conn = get_db()
    row = conn.execute("SELECT * FROM iban_countries WHERE code=?", (code.upper(),)).fetchone()
    conn.close(); return dict(row) if row else None

def db_get_all_iban_countries():
    conn = get_db()
    rows = conn.execute("SELECT * FROM iban_countries ORDER BY name").fetchall()
    conn.close(); return [dict(r) for r in rows]

def db_upsert_iban_country(code, name, length, bban_format, example, structure, notes):
    conn = get_db()
    conn.execute("""INSERT INTO iban_countries (code,name,length,bban_format,example,structure,notes)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(code) DO UPDATE SET name=excluded.name,length=excluded.length,
        bban_format=excluded.bban_format,example=excluded.example,
        structure=excluded.structure,notes=excluded.notes""",
        (code.upper(),name,length,bban_format,example,structure,notes))
    conn.commit(); conn.close()

def db_save_report(entity, entity_type, iban, score, niveau, reco, resume, full_json):
    conn = get_db()
    conn.execute("""INSERT INTO osint_reports
        (entity,entity_type,iban,score,niveau,recommandation,resume,full_json)
        VALUES (?,?,?,?,?,?,?,?)""",
        (entity,entity_type,iban,score,niveau,reco,resume,full_json))
    conn.commit(); conn.close()

def db_get_reports(limit=100):
    conn = get_db()
    rows = conn.execute("SELECT * FROM osint_reports ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
    conn.close(); return [dict(r) for r in rows]

def db_get_watchlist():
    conn = get_db()
    rows = conn.execute("SELECT * FROM watchlist ORDER BY created_at DESC").fetchall()
    conn.close(); return [dict(r) for r in rows]

def db_add_watchlist(entity, entity_type, reason, risk_level, added_by):
    conn = get_db()
    conn.execute("INSERT INTO watchlist (entity,entity_type,reason,risk_level,added_by) VALUES (?,?,?,?,?)",
                 (entity,entity_type,reason,risk_level,added_by))
    conn.commit(); conn.close()

def db_delete_watchlist(wid):
    conn = get_db()
    conn.execute("DELETE FROM watchlist WHERE id=?", (wid,))
    conn.commit(); conn.close()

# Initialise DB on startup
init_db(); seed_banks(); seed_iban_countries()


# ─────────────────────────────────────────────────────────────────
# § 3  IBAN VALIDATION
# ─────────────────────────────────────────────────────────────────

def validate_iban(iban_raw: str) -> dict:
    iban = iban_raw.replace(" ","").replace("-","").upper()
    res = {"raw":iban,"formatted":" ".join(iban[i:i+4] for i in range(0,len(iban),4)),
           "valid":False,"country":"","bank_code":"","branch_code":"",
           "account_no":"","rib_key":"","message":""}
    if len(iban) < 5: res["message"] = "IBAN trop court"; return res
    country = iban[:2]; res["country"] = country
    country_info = db_get_iban_country(country)
    if country_info and len(iban) != country_info["length"]:
        res["message"] = (f"Longueur incorrecte : {len(iban)} chars "
                          f"(attendu {country_info['length']} pour {country_info['name']})")
        return res
    rearranged = iban[4:] + iban[:4]
    numeric = ""
    for ch in rearranged:
        if ch.isdigit(): numeric += ch
        elif ch.isalpha(): numeric += str(ord(ch)-55)
        else: res["message"] = f"Caractère invalide : '{ch}'"; return res
    if int(numeric) % 97 != 1:
        res["message"] = "❌ Clé de contrôle invalide (mod97 échoué)"; return res
    res["valid"] = True; res["message"] = "✅ IBAN valide"
    if country in ("FR","MC") and len(iban) == 27:
        res["bank_code"]=iban[4:9]; res["branch_code"]=iban[9:14]
        res["account_no"]=iban[14:25]; res["rib_key"]=iban[25:27]
    elif country == "DE" and len(iban) == 22: res["bank_code"]=iban[4:12]
    elif country == "GB" and len(iban) == 22: res["bank_code"]=iban[4:8]
    elif country == "NL" and len(iban) == 18: res["bank_code"]=iban[4:8]
    elif country == "ES" and len(iban) == 24: res["bank_code"]=iban[4:8]
    elif country == "IT" and len(iban) == 27: res["bank_code"]=iban[5:10]
    elif country == "CH" and len(iban) == 21: res["bank_code"]=iban[4:9]
    elif country == "BE" and len(iban) == 16: res["bank_code"]=iban[4:7]
    elif country == "LU" and len(iban) == 20: res["bank_code"]=iban[4:7]
    else: res["bank_code"]=iban[4:8]
    return res


# ─────────────────────────────────────────────────────────────────
# § 4  OSINT ENGINE
# ─────────────────────────────────────────────────────────────────

# ── 4a. Query catalogue ───────────────────────────────────────────
# Each entry: (query_template, category, gravity, label)
# {e} is replaced with the entity name at runtime
QUERY_CATALOGUE = [
    # ── Financial crime — FR ──────────────────────────────────────
    ('"{e}" fraude escroquerie arnaque',            "fraud",     "eleve",  "Fraude/Arnaque FR"),
    ('"{e}" corruption pot-de-vin détournement',    "fraud",     "eleve",  "Corruption FR"),
    ('"{e}" blanchiment financement terrorisme',    "sanctions", "eleve",  "Blanchiment FR"),
    ('"{e}" condamné tribunal jugement pénal',      "judicial",  "eleve",  "Condamnation FR"),
    ('"{e}" mis en examen inculpé perquisition',    "judicial",  "moyen",  "Procédure pénale FR"),
    ('"{e}" liquidation judiciaire faillite',       "judicial",  "moyen",  "Faillite FR"),
    ('"{e}" sanction AMF ACPR interdiction',        "sanctions", "eleve",  "Sanction régulateur FR"),
    ('"{e}" liste noire blacklist OFAC',            "sanctions", "eleve",  "Liste sanctions FR"),
    ('"{e}" plainte victime signalement arnaque',   "reputation","moyen",  "Signalement FR"),
    ('"{e}" abus de biens sociaux malversation',    "fraud",     "eleve",  "ABS FR"),
    ('"{e}" fraude fiscale redressement fiscal',    "fraud",     "eleve",  "Fraude fiscale FR"),
    ('"{e}" faux usage de faux falsification',      "fraud",     "moyen",  "Faux FR"),
    ('"{e}" interdit bancaire fichier Banque de France', "judicial","moyen","Interdit bancaire FR"),
    # ── Financial crime — EN ──────────────────────────────────────
    ('"{e}" fraud scam money laundering',           "fraud",     "eleve",  "Fraud EN"),
    ('"{e}" corruption bribery convicted',          "fraud",     "eleve",  "Corruption EN"),
    ('"{e}" sanctions blacklist OFAC banned',       "sanctions", "eleve",  "Sanctions EN"),
    ('"{e}" criminal charges lawsuit arrested',     "judicial",  "eleve",  "Criminal EN"),
    ('"{e}" Ponzi embezzlement misappropriation',   "fraud",     "eleve",  "Embezzlement EN"),
    ('"{e}" regulatory penalty enforcement action', "sanctions", "eleve",  "Regulatory EN"),
    ('"{e}" indicted convicted sentenced prison',   "judicial",  "eleve",  "Sentenced EN"),
    ('"{e}" bankruptcy insolvency default',         "judicial",  "moyen",  "Bankruptcy EN"),
    ('"{e}" scandal controversy warning consumer',  "reputation","moyen",  "Scandal EN"),
    ('"{e}" tax evasion insider trading',           "fraud",     "eleve",  "Tax evasion EN"),
    ('"{e}" terrorist financing links',             "sanctions", "eleve",  "Terror financing EN"),
    # ── Other languages ───────────────────────────────────────────
    ('"{e}" fraude estafa corrupción condenado',    "fraud",     "eleve",  "Fraude ES"),
    ('"{e}" blanqueo dinero sanción investigado',   "sanctions", "eleve",  "Sanciones ES"),
    ('"{e}" Betrug Korruption verurteilt',          "fraud",     "eleve",  "Betrug DE"),
    ('"{e}" Geldwäsche Sanktion Ermittlung',        "sanctions", "eleve",  "Sanktionen DE"),
    ('"{e}" truffa corruzione condannato',          "fraud",     "eleve",  "Truffa IT"),
    ('"{e}" احتيال غسيل أموال فساد',              "fraud",     "eleve",  "Fraude AR"),
    ('"{e}" мошенничество коррупция арест',         "fraud",     "eleve",  "Мошенничество RU"),
    ('"{e}" 欺诈 洗钱 腐败',                        "fraud",     "eleve",  "诈骗 ZH"),
    # ── Official sources ──────────────────────────────────────────
    ('site:opensanctions.org "{e}"',                "sanctions", "eleve",  "OpenSanctions"),
    ('site:ofac.treas.gov "{e}"',                   "sanctions", "eleve",  "OFAC USA"),
    ('site:amf-france.org "{e}"',                   "sanctions", "eleve",  "AMF France"),
    ('site:acpr.banque-france.fr "{e}"',            "sanctions", "eleve",  "ACPR"),
    ('site:sec.gov "{e}"',                          "sanctions", "eleve",  "SEC USA"),
    ('site:legifrance.gouv.fr "{e}"',               "judicial",  "eleve",  "Legifrance"),
    ('site:bodacc.fr "{e}"',                        "judicial",  "moyen",  "BODACC"),
    ('site:infogreffe.fr "{e}"',                    "judicial",  "faible", "Infogreffe"),
    ('site:pappers.fr "{e}"',                       "judicial",  "faible", "Pappers"),
    ('site:societe.com "{e}"',                      "judicial",  "faible", "Societe.com"),
    ('site:companieshouse.gov.uk "{e}"',            "judicial",  "faible", "Companies House UK"),
    ('site:opencorporates.com "{e}"',               "judicial",  "faible", "OpenCorporates"),
    ('"{e}" EU UN OFAC SDN sanctioned',             "sanctions", "eleve",  "SDN sanctions"),
    ('"{e}" Interpol red notice wanted',            "sanctions", "eleve",  "Interpol"),
    ('"{e}" PEP politically exposed person',        "sanctions", "moyen",  "PEP"),
    # ── Scam / review sites ───────────────────────────────────────
    ('site:signal-arnaques.com "{e}"',              "reputation","eleve",  "Signal-arnaques"),
    ('site:cybermalveillance.gouv.fr "{e}"',        "reputation","eleve",  "Cybermalveillance"),
    ('site:trustpilot.com "{e}"',                   "reputation","faible", "Trustpilot"),
    ('"{e}" victime arnaque forum témoignage',      "reputation","moyen",  "Témoignages FR"),
    ('"{e}" scam reported victim forum',            "reputation","moyen",  "Scam reports EN"),
    # ── Social media ──────────────────────────────────────────────
    ('site:twitter.com "{e}" arnaque scam fraud',  "reputation","faible", "Twitter"),
    ('site:reddit.com "{e}" scam fraud arnaque',   "reputation","faible", "Reddit"),
    ('site:facebook.com "{e}" arnaque fraude',     "reputation","faible", "Facebook"),
    # ── Press (world) ─────────────────────────────────────────────
    ('site:lemonde.fr "{e}"',                       "reputation","moyen",  "Le Monde"),
    ('site:lefigaro.fr "{e}"',                      "reputation","moyen",  "Le Figaro"),
    ('site:bfmtv.com "{e}"',                        "reputation","moyen",  "BFM TV"),
    ('site:lesechos.fr "{e}"',                      "reputation","moyen",  "Les Echos"),
    ('site:reuters.com "{e}"',                      "reputation","moyen",  "Reuters"),
    ('site:bloomberg.com "{e}"',                    "reputation","moyen",  "Bloomberg"),
    ('site:theguardian.com "{e}"',                  "reputation","moyen",  "The Guardian"),
    ('site:bbc.com "{e}"',                          "reputation","moyen",  "BBC"),
]

# ── Negative keywords used to identify relevant snippets ──────────
NEGATIVE_KEYWORDS = [
    "fraude","fraud","arnaque","scam","escroquerie","corruption","condamné",
    "convicted","sanctionné","sanctioned","blanchiment","laundering","faillite",
    "bankruptcy","tribunal","lawsuit","arresté","arrested","blacklist","liste noire",
    "détournement","embezzlement","plainte","complaint","victime","victim",
    "jugement","sentenced","peine","prison","inculpé","indicted","perquisition",
    "санкции","мошенничество","احتيال","عقوبات","欺诈","制裁","estafa","corrupción",
    "sanction","interdiction","liquidation","redressement","interdit",
    "poursuivi","poursuites","mis en cause","mis en examen","garde à vue",
    "détournement","abus de biens","malversation","usurpation",
    "signal-arnaques","cybermalveillance","escroc","arnaqueur","suspicious",
]

# ── 4b. Web search ────────────────────────────────────────────────

def search_web(query: str, num: int = 5) -> list:
    """DuckDuckGo HTML search with Bing fallback."""
    from bs4 import BeautifulSoup
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
    results = []
    try:
        r = requests.get(f"https://html.duckduckgo.com/html/?q={quote_plus(query)}",
                         headers={"User-Agent":ua}, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for item in soup.select(".result")[:num]:
            a    = item.find("a", class_="result__a")
            snip = item.find("a", class_="result__snippet")
            if a and str(a.get("href","")).startswith("http"):
                results.append({"title":a.get_text(strip=True),
                                 "url":a.get("href",""),
                                 "snippet":snip.get_text(strip=True) if snip else ""})
    except: pass
    if len(results) < 2:
        try:
            r = requests.get(f"https://www.bing.com/search?q={quote_plus(query)}&count={num}",
                              headers={"User-Agent":ua}, timeout=12)
            soup = BeautifulSoup(r.text, "html.parser")
            for li in soup.select("li.b_algo")[:num]:
                h2=li.find("h2"); a=h2.find("a") if h2 else None; p=li.find("p")
                if a and str(a.get("href","")).startswith("http"):
                    results.append({"title":a.get_text(strip=True),
                                    "url":a.get("href",""),
                                    "snippet":p.get_text(strip=True) if p else ""})
        except: pass
    return results[:num]


def check_opensanctions(name: str) -> dict:
    try:
        r = requests.get(f"https://api.opensanctions.org/search/default?q={quote_plus(name)}&limit=5",
                         timeout=8)
        if r.status_code == 200:
            d = r.json()
            return {"found":d.get("total",0)>0,"count":d.get("total",0),
                    "results":d.get("results",[])[:5]}
    except: pass
    return {"found":False,"count":0,"results":[],"error":"Non disponible"}


def scrape_page(url: str, max_chars=3000) -> str:
    try:
        from bs4 import BeautifulSoup
        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        r = requests.get(url, headers={"User-Agent":ua}, timeout=10, allow_redirects=True)
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script","style","nav","footer","header","aside","iframe"]):
            tag.decompose()
        return soup.get_text(" ", strip=True)[:max_chars]
    except: return ""


# ── 4c. Analysis engine ───────────────────────────────────────────

def run_osint_analysis(entity: str, all_results_with_meta: list,
                       scraped_texts: list, os_result: dict) -> dict:
    """
    Core scoring logic.
    all_results_with_meta: each dict has keys title, url, snippet,
      query_cat, query_gravity, query_label.
    Returns structured analysis dict used by PDF, Excel, and UI.
    """
    entity_low   = entity.lower().strip()
    entity_words = [w for w in entity_low.split() if len(w) >= 3]
    entity_tokens = list(set([entity_low] + entity_words))
    if len(entity_words) >= 2:
        entity_tokens += [entity_words[0], entity_words[-1]]

    def entity_in_text(txt: str) -> bool:
        t = " " + txt.lower() + " "
        return any(tok in t for tok in entity_tokens)

    scores = {k: 0.0 for k in SCORE_WEIGHTS}
    negative_news, all_articles = [], []

    for r in all_results_with_meta:
        title   = r.get("title","")
        url     = r.get("url","")
        snippet = r.get("snippet","")
        domain  = urlparse(url).netloc.replace("www.","")
        cred    = SOURCE_CREDIBILITY.get(domain, 1.0)
        q_cat   = r.get("query_cat","reputation")
        q_grav  = r.get("query_gravity","faible")
        q_label = r.get("query_label","Recherche web")

        combined    = f"{title} {snippet}"
        entity_hit  = entity_in_text(combined) or entity_in_text(url)
        combined_lo = combined.lower()

        # Find negative keywords present in this result
        triggered = [w for w in NEGATIVE_KEYWORDS if w in combined_lo]

        all_articles.append({
            "title":title, "url":url, "snippet":snippet,
            "domain":domain, "entity_mentioned":entity_hit,
            "query_label":q_label, "triggered_keywords":triggered,
        })

        # ── SCORING: result from a targeted negative query IS a signal ──
        # Even if no keyword in snippet, the query already was negative.
        weight = GRAVITY_WEIGHTS.get(q_grav, 1.0) * cred
        scores[q_cat] = scores.get(q_cat, 0.0) + weight

        # Build negative news entry
        nature_map = {
            "sanctions": "Sanction / Liste internationale",
            "fraud":     "Fraude / Crime financier",
            "judicial":  "Procédure judiciaire",
            "reputation":"Signal réputationnel",
            "pep":       "Exposition PEP",
        }
        negative_news.append({
            "titre":        title[:120] or url[:80],
            "source":       domain,
            "url":          url,
            "snippet":      snippet[:200],
            "nature":       nature_map.get(q_cat, q_label),
            "gravite":      q_grav,
            "mots_cles":    triggered[:6] if triggered else [f"[requête: {q_label}]"],
            "query_label":  q_label,
            "score_brut":   round(weight, 1),
            "category":     q_cat,
            "entity_found": entity_hit,
        })

    # OpenSanctions results (highest weight)
    if os_result.get("found"):
        scores["sanctions"] += os_result["count"] * 15
        for r in os_result.get("results",[]):
            negative_news.insert(0, {
                "titre":      f"OpenSanctions : {r.get('caption','')}",
                "source":     "opensanctions.org",
                "url":        f"https://www.opensanctions.org/entities/{r.get('id','')}",
                "snippet":    f"Datasets: {', '.join(r.get('datasets',[]))}",
                "nature":     "Sanction internationale confirmée",
                "gravite":    "eleve",
                "mots_cles":  ["opensanctions"] + r.get("datasets",[])[:3],
                "query_label":"OpenSanctions API",
                "score_brut": 15.0,
                "category":   "sanctions",
                "entity_found":True,
            })

    # Deduplicate by URL (keep highest score per URL)
    seen_urls, neg_dedup = {}, []
    for n in sorted(negative_news, key=lambda x: x["score_brut"], reverse=True):
        u = n.get("url","") or n.get("titre","")[:60]
        if u not in seen_urls:
            seen_urls[u] = True; neg_dedup.append(n)

    # Final risk score (0-100)
    raw = sum(scores.get(cat, 0) * w for cat, w in SCORE_WEIGHTS.items())
    score = min(100, max(0, int(raw * 1.5)))

    niveau = "FAIBLE"
    for lvl, threshold in [("CRITIQUE",70),("ELEVE",40),("MODERE",10)]:
        if score >= threshold: niveau = lvl; break

    reco = ("REFUSER"           if score >= 20 or os_result.get("found") else
            "VIGILANCE_RENFORCEE" if score >= 5 or len(neg_dedup) >= 1 else
            "ACCEPTER")

    nb_entity = sum(1 for n in neg_dedup if n.get("entity_found"))

    # Aggravating / attenuating factors
    aggravants, attenuants = [], []
    if os_result.get("found"):
        aggravants.append(f"{os_result['count']} entrée(s) OpenSanctions confirmée(s)")
    if scores.get("fraud",0) > 3:
        aggravants.append(f"Signaux fraude/corruption (score:{round(scores['fraud'],1)})")
    if scores.get("judicial",0) > 2:
        aggravants.append(f"Signaux procédures judiciaires (score:{round(scores['judicial'],1)})")
    if scores.get("sanctions",0) > 3:
        aggravants.append(f"Signaux sanctions/listes (score:{round(scores['sanctions'],1)})")
    if nb_entity == 0 and len(neg_dedup) > 0:
        attenuants.append(f"{len(neg_dedup)} alertes sans mention directe — possibles faux positifs")
    if not os_result.get("found"):
        attenuants.append("Absent des listes OpenSanctions consultées")

    resume = (
        f"Analyse OSINT de '{entity}' : {len(all_articles)} résultats collectés, "
        f"{len(neg_dedup)} signal(aux) remontés dont {nb_entity} mentionnant "
        f"directement l'entité. Score de risque : {score}/100 ({niveau}). "
        f"Recommandation : {reco}. Revue humaine obligatoire avant toute décision."
    )

    return {
        "score_risque":        score,
        "niveau_risque":       niveau,
        "resume_executif":     resume,
        "recommandation":      reco,
        "negative_news":       neg_dedup,
        "all_articles":        all_articles,
        "sanctions":           {"trouve": os_result.get("found",False) or scores.get("sanctions",0)>3,
                                "details": f"{os_result['count']} OpenSanctions" if os_result.get("found") else ""},
        "litiges_judiciaires": {"trouve": scores.get("judicial",0)>2,
                                "details": f"Score judiciaire : {round(scores.get('judicial',0),1)}"},
        "pep_exposure":        {"trouve": scores.get("pep",0)>2, "details":""},
        "facteurs_aggravants": aggravants,
        "facteurs_attenuants": attenuants,
        "sources_consultees":  list({n["source"] for n in neg_dedup if n.get("source")})[:30],
        "scores_categories":   {k: round(v,1) for k,v in scores.items()},
        "nb_sources_total":    len(all_articles),
        "nb_sources_filtrees": nb_entity,
        "os_result":           os_result,
    }


# ─────────────────────────────────────────────────────────────────
# § 5  PDF REPORT
# ─────────────────────────────────────────────────────────────────

def _pdf_styles():
    """Return ReportLab color palette and paragraph styles."""
    BG   = HexColor("#0a0d14"); SURF = HexColor("#111520"); BORD = HexColor("#1e2535")
    ACC  = HexColor("#00d4ff"); GREEN= HexColor("#00cc66"); RED  = HexColor("#ff3366")
    YEL  = HexColor("#ffcc00"); TEXT = HexColor("#c8d6e5"); MUTED= HexColor("#5a6a7a")
    ACC2 = HexColor("#ff6b35"); WHITE= HexColor("#ffffff")
    styles = getSampleStyleSheet()
    def S(nm,**kw): return ParagraphStyle(nm, parent=styles["Normal"], **kw)
    return {
        "BG":BG,"SURF":SURF,"BORD":BORD,"ACC":ACC,"GREEN":GREEN,"RED":RED,
        "YEL":YEL,"TEXT":TEXT,"MUTED":MUTED,"ACC2":ACC2,"WHITE":WHITE,
        "sH1": S("H1",fontSize=11,textColor=ACC,fontName="Helvetica-Bold",spaceBefore=12,spaceAfter=5),
        "sH2": S("H2",fontSize=9.5,textColor=ACC2,fontName="Helvetica-Bold",spaceBefore=8,spaceAfter=4),
        "sBody":S("B", fontSize=8.5,textColor=TEXT,fontName="Helvetica",spaceAfter=3,leading=13),
        "sSmall":S("Sm",fontSize=7.5,textColor=MUTED,fontName="Helvetica",spaceAfter=2),
        "sWarn": S("W", fontSize=9,textColor=YEL,fontName="Helvetica-Bold",spaceAfter=4),
        "sOK":  S("OK",fontSize=9,textColor=GREEN,fontName="Helvetica-Bold",spaceAfter=4),
        "sRed": S("Rd",fontSize=9,textColor=RED,fontName="Helvetica-Bold",spaceAfter=4),
    }

def generate_osint_pdf(entity: str, analysis: dict, os_result: dict,
                       iban_data: dict = None, bank_data: dict = None,
                       human_decision=None, human_comment="", analyst_name="") -> bytes:
    """
    Generate a professional OSINT-only PDF report.
    Sections:
      1. Cover / Header
      2. Human validation stamp
      3. Executive summary + risk score
      4. Regulatory checks (sanctions, PEP, litigation)
      5. Negative signals table
      6. Full source list for human review
      7. Risk factors
      8. Limits / disclaimer
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=18*mm, rightMargin=18*mm,
                             topMargin=16*mm,  bottomMargin=16*mm)
    c   = _pdf_styles()
    story = []
    now_str = datetime.now().strftime("%d/%m/%Y à %H:%M:%S")

    # ── Helpers ───────────────────────────────────────────────────
    def tbl(rows, widths, style_extra=None):
        t = Table(rows, colWidths=widths)
        ts = TableStyle([
            ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,-1),c["BG"]),
            ("TEXTCOLOR",(0,0),(-1,-1),c["TEXT"]),
            ("GRID",(0,0),(-1,-1),0.3,c["BORD"]),
            ("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ])
        if style_extra:
            for cmd in style_extra: ts.add(*cmd)
        t.setStyle(ts); return t

    score  = analysis.get("score_risque",0)
    niveau = analysis.get("niveau_risque","FAIBLE")
    reco   = analysis.get("recommandation","ACCEPTER")
    neg_n  = analysis.get("negative_news",[])
    all_art= analysis.get("all_articles",[])
    nb_tot = analysis.get("nb_sources_total",0)
    nb_ent = analysis.get("nb_sources_filtrees",0)

    rcolor = {"FAIBLE":c["GREEN"],"MODERE":c["YEL"],"ELEVE":c["ACC2"],"CRITIQUE":c["RED"]}.get(niveau,c["MUTED"])
    rccolor= {"ACCEPTER":c["GREEN"],"VIGILANCE_RENFORCEE":c["YEL"],"REFUSER":c["RED"]}.get(reco,c["MUTED"])

    # ─── 1. HEADER ───────────────────────────────────────────────
    today_header = datetime.now().strftime("%d/%m/%Y")
    story.append(tbl(
        [["🛡️  FinShield OSINT — Rapport de Conformité",
          f"Date : {today_header}\nGénéré le {now_str}"]],
        [120*mm, 52*mm],
        [("BACKGROUND",(0,0),(-1,-1),c["BG"]),
         ("TEXTCOLOR",(0,0),(0,0),c["ACC"]),("TEXTCOLOR",(1,0),(1,0),c["MUTED"]),
         ("FONTNAME",(0,0),(0,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(0,0),14),
         ("FONTSIZE",(1,0),(1,0),8),("ALIGN",(1,0),(1,0),"RIGHT"),
         ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10)]))
    story.append(HRFlowable(width="100%",thickness=1.5,color=c["ACC"],spaceAfter=10))
    story.append(Paragraph(f"Entité analysée : <b>{entity}</b>", c["sH1"]))

    # ─── 2. HUMAN VALIDATION STAMP ───────────────────────────────
    stamp_map = {
        "RAS":             (c["GREEN"], "✅  VALIDATION HUMAINE — RAS (Rien à Signaler)",
                            "L'analyste confirme : aucune information négative retenue après examen."),
        "RISQUE_CONFIRME": (c["RED"],   "⚠️  VALIDATION HUMAINE — RISQUES CONFIRMÉS",
                            "L'analyste confirme les signaux négatifs détectés."),
        None:              (c["MUTED"], "⏳  EN ATTENTE DE VALIDATION HUMAINE",
                            "Résultat automatique — non encore validé par un analyste."),
    }
    s_color, s_text, s_sub = stamp_map.get(human_decision, stamp_map[None])
    story.append(tbl([[s_text]], [172*mm],
        [("BACKGROUND",(0,0),(0,0),c["BG"]),("TEXTCOLOR",(0,0),(0,0),s_color),
         ("FONTNAME",(0,0),(0,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(0,0),10),
         ("ALIGN",(0,0),(0,0),"CENTER"),("TOPPADDING",(0,0),(0,0),10),
         ("BOTTOMPADDING",(0,0),(0,0),6),("BOX",(0,0),(0,0),1.5,s_color)]))
    story.append(Paragraph(s_sub, c["sSmall"]))
    if analyst_name or human_comment:
        rows = []
        if analyst_name:    rows.append(["Analyste", analyst_name])
        if human_comment:   rows.append(["Commentaire", human_comment])
        rows.append(["Date", now_str])
        story.append(tbl(rows, [38*mm,134*mm],
            [("TEXTCOLOR",(0,0),(0,-1),c["MUTED"])]))
    story.append(Spacer(1,8))

    # ─── 3. SCORE BOX ────────────────────────────────────────────
    story.append(tbl(
        [[f"Score : {score}/100", f"Niveau : {niveau}",
          f"Recommandation : {reco}", f"Sources : {nb_tot}"]],
        [40*mm,38*mm,60*mm,34*mm],
        [("BACKGROUND",(0,0),(-1,-1),c["SURF"]),
         ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8.5),
         ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
         ("TOPPADDING",(0,0),(-1,-1),9),("BOTTOMPADDING",(0,0),(-1,-1),9),
         ("TEXTCOLOR",(0,0),(0,0),c["TEXT"]),
         ("TEXTCOLOR",(1,0),(1,0),rcolor),("TEXTCOLOR",(2,0),(2,0),rccolor),
         ("TEXTCOLOR",(3,0),(3,0),c["MUTED"]),
         ("BOX",(0,0),(-1,-1),0.5,c["BORD"]),("GRID",(0,0),(-1,-1),0.5,c["BORD"])]))
    story.append(Spacer(1,6))

    # ─── 4. RÉSUMÉ EXÉCUTIF ──────────────────────────────────────
    story.append(Paragraph("1. Résumé Exécutif", c["sH1"]))
    story.append(Paragraph(analysis.get("resume_executif","—"), c["sBody"]))
    story.append(HRFlowable(width="100%",thickness=0.5,color=c["BORD"],spaceAfter=6))

    # ─── 5. IBAN (if available) ───────────────────────────────────
    if iban_data and iban_data.get("raw"):
        story.append(Paragraph("2. Vérification IBAN", c["sH1"]))
        rows = [["Champ","Valeur"]]
        for f,v in [("IBAN",iban_data.get("formatted","")),
                    ("Statut",iban_data.get("message","")),
                    ("Pays",iban_data.get("country","")),
                    ("Code banque",iban_data.get("bank_code",""))]:
            if v: rows.append([f,v])
        if bank_data:
            rows += [["Banque",bank_data.get("name","")],
                     ["Adresse",f"{bank_data.get('address','')} {bank_data.get('city','')}".strip()],
                     ["BIC",bank_data.get("bic","")],
                     ["Type",bank_data.get("type","")]]
        story.append(tbl(rows, [40*mm,132*mm],
            [("BACKGROUND",(0,0),(-1,0),c["SURF"]),("TEXTCOLOR",(0,0),(-1,0),c["ACC"]),
             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
             ("TEXTCOLOR",(0,1),(0,-1),c["MUTED"]),
             ("ROWBACKGROUNDS",(0,1),(-1,-1),[c["BG"],c["SURF"]])]))
        story.append(Spacer(1,8))

    # ─── 6. REGULATORY CHECKS ────────────────────────────────────
    story.append(Paragraph("3. Vérifications Réglementaires (KYC/AML)", c["sH1"]))
    san = analysis.get("sanctions",{})
    lit = analysis.get("litiges_judiciaires",{})
    pep = analysis.get("pep_exposure",{})
    reg_rows = [
        ["Catégorie","Statut","Détails"],
        ["Sanctions / Listes",
         "DÉTECTÉ" if san.get("trouve") else "Non détecté",
         san.get("details","")[:85]],
        ["OpenSanctions",
         f"{os_result.get('count',0)} hit(s)",
         ", ".join(r.get("caption","") for r in os_result.get("results",[])[:2])[:85]],
        ["Litiges judiciaires",
         "DÉTECTÉ" if lit.get("trouve") else "Non détecté",
         lit.get("details","")[:85]],
        ["Exposition PEP",
         "DÉTECTÉ" if pep.get("trouve") else "Non détecté",
         pep.get("details","")[:85]],
    ]
    # ── build dynamic style for regulatory table (fix _tStyle) ──
    reg_style_cmds = [
        ("BACKGROUND",(0,0),(-1,0),c["SURF"]),
        ("TEXTCOLOR",(0,0),(-1,0),c["ACC"]),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[c["BG"],c["SURF"]]),
    ]
    for i in range(1, len(reg_rows)):
        col = c["RED"] if "DÉTECTÉ" in str(reg_rows[i][1]) or (i==2 and os_result.get("count",0)>0) else c["GREEN"]
        reg_style_cmds.append(("TEXTCOLOR",(1,i),(1,i),col))
        reg_style_cmds.append(("FONTNAME",(1,i),(1,i),"Helvetica-Bold"))
    story.append(tbl(reg_rows, [45*mm,28*mm,99*mm], reg_style_cmds))
    story.append(Spacer(1,8))

    # ─── 7. NEGATIVE SIGNALS TABLE ───────────────────────────────
    story.append(Paragraph(f"4. Signaux Négatifs Détectés ({len(neg_n)})", c["sH1"]))
    story.append(Paragraph(
        "Chaque ligne ci-dessous provient d'une requête ciblée sur un thème négatif "
        "(fraude, sanctions, litiges…). La colonne 'Entité ✓' indique si le nom "
        "recherché apparaît directement dans le titre ou l'extrait.", c["sSmall"]))
    story.append(Spacer(1,4))
    if neg_n:
        gmap = {"faible":c["GREEN"],"moyen":c["YEL"],"eleve":c["RED"]}
        nh = [["Titre / Source","Nature","Gravité","Mots-clés","Entité ✓"]]
        for n in neg_n[:20]:
            nh.append([
                f"{n.get('titre','')[:55]}\n{n.get('source','')[:30]}",
                n.get("nature","")[:28],
                n.get("gravite","").upper(),
                ", ".join(n.get("mots_cles",[])[:3])[:30],
                "✓" if n.get("entity_found") else "—",
            ])
        # ── build dynamic style (fix _tStyle) ──
        neg_style_cmds = [
            ("BACKGROUND",(0,0),(-1,0),c["SURF"]),
            ("TEXTCOLOR",(0,0),(-1,0),c["ACC"]),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[c["BG"],c["SURF"]]),
        ]
        for i, n in enumerate(neg_n[:20], 1):
            col = gmap.get(n.get("gravite","").lower(), c["MUTED"])
            neg_style_cmds.append(("TEXTCOLOR",(2,i),(2,i),col))
            neg_style_cmds.append(("FONTNAME",(2,i),(2,i),"Helvetica-Bold"))
            if n.get("entity_found"):
                neg_style_cmds.append(("TEXTCOLOR",(4,i),(4,i),c["GREEN"]))
        story.append(tbl(nh, [72*mm,35*mm,18*mm,32*mm,15*mm], neg_style_cmds))
    else:
        story.append(Paragraph("Aucun signal négatif détecté.", c["sBody"]))
    story.append(Spacer(1,6))

    # ─── 8. FULL SOURCE LIST FOR HUMAN REVIEW ────────────────────
    story.append(PageBreak())
    story.append(Paragraph(f"5. Sources Complètes pour Revue Humaine ({len(all_art)} articles)", c["sH1"]))
    story.append(Paragraph(
        "Liste exhaustive des pages collectées. L'analyste doit consulter chaque "
        "article avant toute décision. La colonne 'Entité ✓' confirme la présence "
        "du nom dans le contenu visible.", c["sSmall"]))
    story.append(Spacer(1,4))
    if all_art:
        ah = [["Titre","Domaine","Requête thématique","Entité ✓"]]
        for a in all_art[:50]:
            ah.append([
                a.get("title","")[:60],
                a.get("domain","")[:25],
                a.get("query_label","")[:28],
                "✓" if a.get("entity_mentioned") else "—",
            ])
        # ── build dynamic style (fix _tStyle) ──
        src_style_cmds = [
            ("BACKGROUND",(0,0),(-1,0),c["SURF"]),
            ("TEXTCOLOR",(0,0),(-1,0),c["ACC"]),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[c["BG"],c["SURF"]]),
        ]
        for i, a in enumerate(all_art[:50], 1):
            if a.get("entity_mentioned"):
                src_style_cmds.append(("TEXTCOLOR",(3,i),(3,i),c["GREEN"]))
        story.append(tbl(ah, [80*mm,32*mm,42*mm,18*mm], src_style_cmds))
    story.append(Spacer(1,8))

    # ─── 9. RISK FACTORS ─────────────────────────────────────────
    fa = analysis.get("facteurs_aggravants",[])
    fat = analysis.get("facteurs_attenuants",[])
    if fa or fat:
        story.append(Paragraph("6. Facteurs de Risque", c["sH1"]))
        c1 = [Paragraph("⚠ Aggravants", ParagraphStyle("fa",fontSize=8.5,
              textColor=c["RED"],fontName="Helvetica-Bold"))]
        c1 += [Paragraph(f"• {f}", c["sBody"]) for f in fa] or [Paragraph("Aucun", c["sBody"])]
        c2 = [Paragraph("✓ Atténuants", ParagraphStyle("ft",fontSize=8.5,
              textColor=c["GREEN"],fontName="Helvetica-Bold"))]
        c2 += [Paragraph(f"• {f}", c["sBody"]) for f in fat] or [Paragraph("Aucun", c["sBody"])]
        story.append(Table([[c1,c2]], colWidths=[86*mm,86*mm],
            style=TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
                               ("BACKGROUND",(0,0),(-1,-1),c["BG"]),
                               ("GRID",(0,0),(-1,-1),0.3,c["BORD"]),
                               ("TOPPADDING",(0,0),(-1,-1),7),
                               ("LEFTPADDING",(0,0),(-1,-1),7)])))
        story.append(Spacer(1,8))

    # ─── 10. LIMITS / DISCLAIMER ─────────────────────────────────
    story.append(Paragraph("7. Limites et Remarques", c["sH1"]))
    disclaimer = (
        "Cette analyse est produite à partir de sources publiques (OSINT) disponibles "
        "à la date de génération du rapport. Elle ne constitue pas une preuve juridique "
        "ni un avis de droit. Les résultats dépendent de la qualité et de l'exhaustivité "
        "des sources interrogées, ainsi que de la présence de l'entité dans ces sources. "
        "Des homonymes peuvent générer des faux positifs. Une validation humaine par un "
        "analyste qualifié est indispensable avant toute prise de décision. "
        "Ce document est à usage interne exclusivement et ne peut être transmis à des tiers "
        "sans autorisation préalable de la direction conformité."
    )
    story.append(Paragraph(disclaimer, c["sSmall"]))

    # ─── FOOTER ──────────────────────────────────────────────────
    story.append(HRFlowable(width="100%",thickness=0.5,color=c["BORD"],spaceAfter=4))
    verdict = f"VALIDATION : {human_decision or 'EN ATTENTE'}"
    if analyst_name: verdict += f" par {analyst_name}"

    # Footer row: copyright left | verdict center | date right
    today_str = datetime.now().strftime("%d/%m/%Y")
    _sFooter = ParagraphStyle("footer", fontSize=6, textColor=c["MUTED"],
                               fontName="Helvetica", leading=8)
    footer_tbl = Table(
        [[
            Paragraph(f"© {datetime.now().year} Mackenson CINEUS · Tous droits réservés", _sFooter),
            Paragraph(f"FinShield OSINT v3 · {verdict} · CONFIDENTIEL", _sFooter),
            Paragraph(f"Généré le {today_str}", _sFooter),
        ]],
        colWidths=[60*mm, 80*mm, 32*mm],
        style=TableStyle([
            ("FONTSIZE",(0,0),(-1,-1),6),
            ("TEXTCOLOR",(0,0),(-1,-1),c["MUTED"]),
            ("ALIGN",(0,0),(0,0),"LEFT"),
            ("ALIGN",(1,0),(1,0),"CENTER"),
            ("ALIGN",(2,0),(2,0),"RIGHT"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),3),
            ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ])
    )
    story.append(footer_tbl)

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
# § 6  EXCEL HISTORY
# ─────────────────────────────────────────────────────────────────

def _xl_header_style(cell, bg="#0a0d14", fg="00D4FF", bold=True):
    cell.font   = Font(bold=bold, color=fg, name="Arial", size=9)
    cell.fill   = PatternFill("solid", start_color=bg.lstrip("#"), end_color=bg.lstrip("#"))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _xl_cell_style(cell, fg="C8D6E5", bg="111520", bold=False, align="left"):
    cell.font   = Font(bold=bold, color=fg, name="Arial", size=8)
    cell.fill   = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def _xl_border():
    thin = Side(style="thin", color="1e2535")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def append_to_excel_history(entity: str, entity_type: str, analysis: dict,
                             os_result: dict, iban_data: dict = None,
                             analyst_name: str = "", human_decision=None) -> str:
    """
    Append a search result row to the Excel history file.
    Creates the file with proper structure if it does not exist.
    Returns the file path.
    """
    path = EXCEL_HISTORY_PATH

    # ── Sheet 1: Historique ───────────────────────────────────────
    HIST_HEADERS = [
        "Date", "Entité", "Type", "Score", "Niveau de risque",
        "Recommandation", "Nb alertes", "Nb entité directe",
        "Nb sources total", "Sanctions OpenSanctions",
        "Listes sanctions", "Litiges", "PEP",
        "IBAN", "Pays IBAN", "Banque",
        "Analyste", "Décision humaine", "Résumé",
    ]

    # ── Sheet 2: Alertes ──────────────────────────────────────────
    ALERT_HEADERS = [
        "Date", "Entité", "Titre alerte", "Source", "Nature",
        "Gravité", "Mots-clés", "Entité mentionnée", "URL",
    ]

    # ── Sheet 3: Sources ─────────────────────────────────────────
    SOURCE_HEADERS = [
        "Date", "Entité", "Domaine", "Titre", "Requête thématique",
        "Entité mentionnée", "URL",
    ]

    # Load or create workbook
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    def get_or_create_sheet(name, headers):
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                _xl_header_style(cell)
                ws.column_dimensions[get_column_letter(col)].width = max(12, len(h)+2)
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"
        return wb[name]

    ws_hist = get_or_create_sheet("Historique", HIST_HEADERS)
    ws_alert= get_or_create_sheet("Alertes",   ALERT_HEADERS)
    ws_src  = get_or_create_sheet("Sources",   SOURCE_HEADERS)

    now_str    = datetime.now().strftime("%d/%m/%Y %H:%M")
    score      = analysis.get("score_risque", 0)
    niveau     = analysis.get("niveau_risque","FAIBLE")
    reco       = analysis.get("recommandation","ACCEPTER")
    neg_n      = analysis.get("negative_news",[])
    nb_entity  = analysis.get("nb_sources_filtrees",0)
    nb_tot     = analysis.get("nb_sources_total",0)
    san        = analysis.get("sanctions",{})
    lit        = analysis.get("litiges_judiciaires",{})
    pep        = analysis.get("pep_exposure",{})
    iban_str   = iban_data.get("formatted","") if iban_data else ""
    pays_iban  = iban_data.get("country","")   if iban_data else ""
    banque_str = ""

    # Risk color for Score cell
    score_bg = {"FAIBLE":"00cc66","MODERE":"ffcc00","ELEVE":"ff6b35","CRITIQUE":"ff3366"}.get(niveau,"5a6a7a")

    # Row for Historique
    hist_row = [
        now_str, entity, entity_type, score, niveau,
        reco, len(neg_n), nb_entity, nb_tot,
        os_result.get("count",0),
        "OUI" if san.get("trouve") else "NON",
        "OUI" if lit.get("trouve") else "NON",
        "OUI" if pep.get("trouve") else "NON",
        iban_str, pays_iban, banque_str,
        analyst_name or "—",
        human_decision or "EN ATTENTE",
        analysis.get("resume_executif","")[:200],
    ]
    next_row = ws_hist.max_row + 1
    for col, val in enumerate(hist_row, 1):
        cell = ws_hist.cell(row=next_row, column=col, value=val)
        bg = "111520" if next_row % 2 == 0 else "0a0d14"
        _xl_cell_style(cell, bg=bg)
        cell.border = _xl_border()
    # Color the score cell
    score_cell = ws_hist.cell(row=next_row, column=4)
    score_cell.fill = PatternFill("solid", start_color=score_bg, end_color=score_bg)
    score_cell.font = Font(bold=True, color="0a0d14", name="Arial", size=8)
    # Color niveau cell
    niveau_cell = ws_hist.cell(row=next_row, column=5)
    niveau_cell.fill = PatternFill("solid", start_color=score_bg, end_color=score_bg)
    niveau_cell.font = Font(bold=True, color="0a0d14", name="Arial", size=8)

    # Rows for Alertes
    for n in neg_n[:50]:
        alert_row = [
            now_str, entity,
            n.get("titre","")[:100], n.get("source",""),
            n.get("nature",""), n.get("gravite","").upper(),
            ", ".join(n.get("mots_cles",[])[:5]),
            "OUI" if n.get("entity_found") else "NON",
            n.get("url","")[:150],
        ]
        next_a = ws_alert.max_row + 1
        for col, val in enumerate(alert_row, 1):
            cell = ws_alert.cell(row=next_a, column=col, value=val)
            bg = "111520" if next_a % 2 == 0 else "0a0d14"
            _xl_cell_style(cell, bg=bg)
            cell.border = _xl_border()

    # Rows for Sources
    for a in analysis.get("all_articles",[])[:100]:
        src_row = [
            now_str, entity,
            a.get("domain",""), a.get("title","")[:80],
            a.get("query_label",""),
            "OUI" if a.get("entity_mentioned") else "NON",
            a.get("url","")[:150],
        ]
        next_s = ws_src.max_row + 1
        for col, val in enumerate(src_row, 1):
            cell = ws_src.cell(row=next_s, column=col, value=val)
            bg = "111520" if next_s % 2 == 0 else "0a0d14"
            _xl_cell_style(cell, bg=bg)
            cell.border = _xl_border()

    # Auto-fit columns (approximate)
    for ws in [ws_hist, ws_alert, ws_src]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(60, max(10, max_len+2))

    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────
# § 7  STREAMLIT SIDEBAR
# ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style='font-family:IBM Plex Mono,monospace;color:#00d4ff;font-size:1.1rem;margin-bottom:4px;'>
    🛡️ FinShield OSINT v3</div>
    <div style='font-size:0.7rem;color:#5a6a7a;letter-spacing:2px;margin-bottom:20px;'>
    COMPLIANCE · INTELLIGENCE · KYC/AML</div>
    """, unsafe_allow_html=True)

    with st.expander("⚡ LLM Boost (optionnel)"):
        api_key = st.text_input("Clé Groq (gratuite)", type="password",
                                 placeholder="gsk_...", key="groq_key")
        st.markdown("""<div style='font-size:0.72rem;color:#5a6a7a;'>
        Clé gratuite : <a href='https://console.groq.com' target='_blank'
        style='color:#00d4ff;'>console.groq.com</a></div>""", unsafe_allow_html=True)

    st.markdown("---")
    conn = get_db()
    nb_banks   = conn.execute("SELECT COUNT(*) FROM banks").fetchone()[0]
    nb_ibanc   = conn.execute("SELECT COUNT(*) FROM iban_countries").fetchone()[0]
    nb_reports = conn.execute("SELECT COUNT(*) FROM osint_reports").fetchone()[0]
    nb_watch   = conn.execute("SELECT COUNT(*) FROM watchlist").fetchone()[0]
    conn.close()
    st.markdown(f"""<div style='font-size:0.72rem;color:#5a6a7a;line-height:2.0;'>
    <b style='color:#c8d6e5;'>Base de données</b><br>
    🏦 <span style='color:#00d4ff;'>{nb_banks}</span> banques indexées<br>
    🌍 <span style='color:#00d4ff;'>{nb_ibanc}</span> pays IBAN<br>
    📋 <span style='color:#00d4ff;'>{nb_reports}</span> rapports générés<br>
    👁 <span style='color:#00d4ff;'>{nb_watch}</span> entités sous surveillance<br>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.caption("SQLite local · Données persistantes")


# ─────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class='header-strip'>
  <span style='font-size:2.5rem;'>🛡️</span>
  <div>
    <div class='app-title'>FinShield OSINT</div>
    <div class='app-sub'>Conformité · Due Diligence · Intelligence Financière v3</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🏦  VÉRIFICATION IBAN",
    "🔍  ANALYSE OSINT",
    "📋  RECHERCHE BANQUE",
    "🗄️  GESTION BASE",
    "📁  HISTORIQUE & SURVEILLANCE",
])


# ─────────────────────────────────────────────────────────────────
# § 8  TAB 1 — IBAN VERIFICATION
# ─────────────────────────────────────────────────────────────────
with tab1:
    st.markdown("## Vérification IBAN")
    st.markdown("<div class='info-box'>43 pays supportés (ISO 13616). Validation mod-97, décomposition complète, identification banque via base locale.</div>", unsafe_allow_html=True)

    col1, col2 = st.columns([3,1])
    with col1:
        iban_input = st.text_input("IBAN", placeholder="FR76 3000 4000 0000 0000 0000 000")
    with col2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("▶ VÉRIFIER", key="btn_iban"):
            res = validate_iban(iban_input)
            country_info = db_get_iban_country(res["country"]) if res["country"] else None
            if res["valid"]:
                st.markdown(f"<div class='ok-box'>{res['message']}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='danger-box'>{res['message']}</div>", unsafe_allow_html=True)
            ca, cb, cc = st.columns([1.2,1,1])
            with ca:
                st.markdown("#### Décomposition")
                fields = [("IBAN formaté",res.get("formatted","")),
                          ("Pays",f"{res.get('country','')} — {country_info['name'] if country_info else 'Inconnu'}"),
                          ("Code banque",res.get("bank_code","")),
                          ("Code guichet",res.get("branch_code","")),
                          ("N° compte",res.get("account_no","")),
                          ("Clé RIB",res.get("rib_key",""))]
                for label, val in fields:
                    if val and val.strip():
                        st.markdown(f"""<div class='result-row'>
                        <span style='color:#5a6a7a;font-size:0.72rem;text-transform:uppercase;'>{label}</span><br>
                        <span style='font-family:IBM Plex Mono,monospace;'>{val}</span></div>""", unsafe_allow_html=True)
            with cb:
                st.markdown("#### Structure pays")
                if country_info:
                    st.markdown(f"""<div class='metric-card'>
                    <div class='label'>Pays</div><div class='value' style='font-size:1rem;'>
                    {country_info['name']} ({country_info['code']})</div>
                    <div class='sub'>Longueur : {country_info['length']} chars</div></div>""", unsafe_allow_html=True)
                    if country_info.get("example"):
                        st.markdown(f"<div class='info-box'>Exemple : <code>{country_info['example']}</code></div>", unsafe_allow_html=True)
            with cc:
                st.markdown("#### Banque")
                if res.get("bank_code"):
                    bank = db_get_bank_by_code(res["bank_code"])
                    if bank:
                        st.markdown(f"""<div class='metric-card'>
                        <div class='label'>Établissement</div>
                        <div class='value' style='font-size:0.9rem;'>{bank['name']}</div>
                        <div class='sub'>{bank.get('city','')} · {bank.get('bic','')}</div>
                        <div style='margin-top:6px;'><span class='badge-low'>{bank.get('type','')}</span></div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        st.markdown(f"<div class='warn-box'>Code <b>{res['bank_code']}</b> non trouvé</div>", unsafe_allow_html=True)

    with st.expander("🌍 Référentiel IBAN"):
        all_countries = db_get_all_iban_countries()
        df_c = pd.DataFrame(all_countries)[["code","name","length","example"]]
        df_c.columns = ["Code","Pays","Longueur","Exemple"]
        st.dataframe(df_c, use_container_width=True, height=400)


# ─────────────────────────────────────────────────────────────────
# § 9  TAB 2 — OSINT ANALYSIS
# ─────────────────────────────────────────────────────────────────
with tab2:
    st.markdown("## Screening OSINT & Due Diligence")
    st.markdown(f"""<div class='info-box'>
    <b>Moteur v3 — {len(QUERY_CATALOGUE)} requêtes thématiques</b> · 
    Fraude, blanchiment, sanctions, PEP, litiges, presse mondiale, avis, réseaux sociaux.
    Chaque résultat issu d'une requête négative est automatiquement qualifié comme signal.
    <b>Revue humaine obligatoire</b> avant toute décision.
    </div>""", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([3,1,1])
    with c1:
        entity_input = st.text_input("Entité à analyser",
                                      placeholder="Ex: Jean Dupont  ou  Société XYZ SAS", key="entity_osint")
    with c2:
        entity_type = st.selectbox("Type", ["Entreprise","Personne physique","Groupe bancaire","Autre"])
    with c3:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        launch_btn = st.button("▶ LANCER LE SCREENING", key="btn_osint")

    with st.expander("⚙️ Options avancées"):
        oa, ob = st.columns(2)
        with oa:
            linked_iban  = st.text_input("IBAN lié (optionnel)", key="linked_iban")
            add_to_watch = st.checkbox("Ajouter à la surveillance après analyse")
        with ob:
            groq_key_tab = st.text_input("Clé Groq (optionnel)", type="password",
                                          placeholder="gsk_...", key="groq_tab")
            nb_results_per_query = st.slider("Résultats par requête", 3, 8, 5,
                                              help="Plus = plus exhaustif mais plus lent")

    # Session state
    for k,v in [("osint_analysis",None),("osint_entity",""),("osint_iban_data",{}),
                ("osint_bank_data",{}),("osint_os_result",{}),("osint_has_risk",False),
                ("osint_all_results",[])]:
        if k not in st.session_state: st.session_state[k] = v

    if launch_btn and entity_input.strip():
        entity = entity_input.strip()
        prog = st.progress(0); stat = st.empty()

        # STEP 1 — OpenSanctions
        stat.markdown("🔎 **[1/4]** OpenSanctions…")
        os_result = check_opensanctions(entity)
        prog.progress(5)

        # STEP 2 — Run all queries
        stat.markdown(f"🌐 **[2/4]** {len(QUERY_CATALOGUE)} requêtes thématiques…")
        all_results_with_meta = []
        for i, (q_tpl, q_cat, q_grav, q_label) in enumerate(QUERY_CATALOGUE):
            q_str = q_tpl.format(e=entity)
            hits  = search_web(q_str, num=nb_results_per_query)
            for h in hits:
                h["query_cat"]     = q_cat
                h["query_gravity"] = q_grav
                h["query_label"]   = q_label
            all_results_with_meta.extend(hits)
            prog.progress(5 + int((i+1)/len(QUERY_CATALOGUE)*55))
            time.sleep(0.06)

        # Deduplicate by URL (keep first occurrence)
        seen, unique = set(), []
        for r in all_results_with_meta:
            u = r.get("url","")
            if u and u not in seen: seen.add(u); unique.append(r)
        all_results_with_meta = unique

        # STEP 3 — IBAN
        stat.markdown("🏦 **[3/4]** Vérification IBAN…")
        iban_data, bank_data = {}, {}
        if linked_iban.strip():
            iban_data = validate_iban(linked_iban.strip())
            if iban_data.get("bank_code"):
                b = db_get_bank_by_code(iban_data["bank_code"])
                if b: bank_data = b
        prog.progress(70)

        # STEP 4 — Analysis
        stat.markdown("🔍 **[4/4]** Analyse et scoring…")
        analysis = run_osint_analysis(entity, all_results_with_meta, [], os_result)
        prog.progress(90)

        # Save to DB
        db_save_report(entity, entity_type, linked_iban or "",
                       analysis["score_risque"], analysis["niveau_risque"],
                       analysis["recommandation"], analysis["resume_executif"],
                       json.dumps(analysis, ensure_ascii=False))

        # Save to Excel history
        try:
            append_to_excel_history(entity, entity_type, analysis, os_result,
                                     iban_data if iban_data.get("raw") else None)
        except Exception as ex:
            st.warning(f"Excel history : {ex}")

        if add_to_watch:
            db_add_watchlist(entity, entity_type, "Screening auto",
                             analysis["niveau_risque"], "Auto")

        # Persist
        score, niveau, reco = (analysis["score_risque"], analysis["niveau_risque"],
                                analysis["recommandation"])
        has_risk = (score >= 5 or os_result.get("found")
                    or analysis["sanctions"]["trouve"]
                    or analysis["litiges_judiciaires"]["trouve"]
                    or len(analysis["negative_news"]) >= 1
                    or niveau in ("MODERE","ELEVE","CRITIQUE"))
        st.session_state.update({
            "osint_analysis":    analysis,
            "osint_entity":      entity,
            "osint_iban_data":   iban_data,
            "osint_bank_data":   bank_data,
            "osint_os_result":   os_result,
            "osint_has_risk":    has_risk,
            "osint_all_results": all_results_with_meta,
        })
        prog.progress(100); stat.empty()

    # ── DISPLAY ───────────────────────────────────────────────────
    analysis    = st.session_state["osint_analysis"]
    entity_d    = st.session_state["osint_entity"]
    iban_data   = st.session_state["osint_iban_data"]
    bank_data   = st.session_state["osint_bank_data"]
    os_result   = st.session_state["osint_os_result"]
    has_risk    = st.session_state["osint_has_risk"]

    if analysis:
        st.markdown("---")
        score  = analysis["score_risque"]
        niveau = analysis["niveau_risque"]
        reco   = analysis["recommandation"]
        neg_n  = analysis["negative_news"]
        nb_tot = analysis["nb_sources_total"]
        nb_ent = analysis["nb_sources_filtrees"]

        # ── RAS banner ────────────────────────────────────────────
        if not has_risk:
            st.markdown(f"""<div style='background:rgba(0,255,136,0.07);border:2px solid #00ff88;
            border-radius:8px;padding:28px;text-align:center;'>
            <div style='font-size:2.5rem;'>✅</div>
            <div style='font-family:IBM Plex Mono,monospace;font-size:1.3rem;color:#00ff88;margin:10px 0;'>
            RAS — AUCUN RISQUE DÉTECTÉ</div>
            <div style='color:#c8d6e5;'><b>{entity_d}</b></div>
            <div style='color:#5a6a7a;font-size:0.8rem;margin-top:10px;'>
            Score : {score}/100 · {nb_tot} sources · {nb_ent} mentions directes</div>
            </div>""", unsafe_allow_html=True)
            st.markdown(f"<div class='ok-box'>{analysis['resume_executif']}</div>", unsafe_allow_html=True)

        else:
            # Score metrics
            bmap  = {"FAIBLE":"badge-low","MODERE":"badge-medium","ELEVE":"badge-high","CRITIQUE":"badge-high"}.get(niveau,"badge-medium")
            rc_c  = {"ACCEPTER":"#00ff88","VIGILANCE_RENFORCEE":"#ffcc00","REFUSER":"#ff3366"}.get(reco,"#5a6a7a")
            mc1,mc2,mc3,mc4 = st.columns(4)
            mc1.markdown(f"<div class='metric-card'><div class='label'>Score</div><div class='value'>{score}<span style='font-size:0.8rem;color:#5a6a7a;'>/100</span></div></div>", unsafe_allow_html=True)
            mc2.markdown(f"<div class='metric-card'><div class='label'>Niveau</div><div style='margin-top:12px;'><span class='{bmap}'>{niveau}</span></div></div>", unsafe_allow_html=True)
            sc_t = f"🔴 {os_result.get('count',0)} hit(s)" if os_result.get("found") else "✅ Aucune sanction"
            sc_c = "#ff3366" if os_result.get("found") else "#00ff88"
            mc3.markdown(f"<div class='metric-card'><div class='label'>Sanctions</div><div class='value' style='font-size:0.8rem;color:{sc_c};'>{sc_t}</div></div>", unsafe_allow_html=True)
            mc4.markdown(f"<div class='metric-card'><div class='label'>Recommandation</div><div class='value' style='font-size:0.72rem;color:{rc_c};'>{reco}</div></div>", unsafe_allow_html=True)

            st.markdown(f"<div class='warn-box'><b>⚠ Résumé :</b> {analysis['resume_executif']}</div>", unsafe_allow_html=True)

            dl, dr = st.columns(2)
            with dl:
                st.markdown(f"#### 📰 Signaux détectés ({len(neg_n)})")
                for n in neg_n[:15]:
                    g   = n.get("gravite","").lower()
                    cls = {"faible":"info-box","moyen":"warn-box","eleve":"danger-box"}.get(g,"warn-box")
                    url = n.get("url","")
                    lnk = f" <a href='{url}' target='_blank' style='color:#00d4ff;font-size:0.72rem;'>→ lire</a>" if url else ""
                    kws = ", ".join(n.get("mots_cles",[])[:4])
                    direct = " <span style='color:#00ff88;font-size:0.7rem;'>✓ direct</span>" if n.get("entity_found") else " <span style='color:#ffcc00;font-size:0.7rem;'>⚠ vérifier</span>"
                    st.markdown(f"""<div class='{cls}'>
                    <b>{n.get('titre','')[:100]}</b>{lnk}{direct}<br>
                    <small style='color:#5a6a7a;'>{n.get('source','')} · <b style='color:#c8d6e5;'>{n.get('nature','')}</b></small><br>
                    <small style='color:#ffcc00;'>🔑 {kws}</small>
                    </div>""", unsafe_allow_html=True)

            with dr:
                st.markdown("#### 🚨 Sanctions & PEP")
                if os_result.get("found"):
                    st.markdown(f"<div class='danger-box'>🔴 {os_result['count']} entrée(s) OpenSanctions</div>", unsafe_allow_html=True)
                    for r in os_result.get("results",[])[:3]:
                        st.markdown(f"<div class='result-row'><b>{r.get('caption','')}</b> · {', '.join(r.get('datasets',[]))}</div>", unsafe_allow_html=True)
                else:
                    st.markdown("<div class='ok-box'>Absent des listes de sanctions</div>", unsafe_allow_html=True)
                pep = analysis.get("pep_exposure",{})
                if pep.get("trouve"):
                    st.markdown(f"<div class='warn-box'>⚠️ Indicateurs PEP détectés</div>", unsafe_allow_html=True)
                for f in analysis.get("facteurs_aggravants",[]):
                    st.markdown(f"<div class='danger-box'>🔺 {f}</div>", unsafe_allow_html=True)
                for f in analysis.get("facteurs_attenuants",[]):
                    st.markdown(f"<div class='ok-box'>🔻 {f}</div>", unsafe_allow_html=True)

        # All articles
        all_art = analysis.get("all_articles",[])
        with st.expander(f"📋 Toutes les sources collectées ({len(all_art)} articles)"):
            art_direct  = [a for a in all_art if a.get("entity_mentioned")]
            art_other   = [a for a in all_art if not a.get("entity_mentioned")]
            st.markdown(f"**Mentions directes de '{entity_d}' ({len(art_direct)})**")
            for a in art_direct:
                url = a.get("url","")
                lnk = f"<a href='{url}' target='_blank' style='color:#00d4ff;'>→ lire</a>" if url else ""
                st.markdown(f"""<div class='result-row' style='border-left:3px solid #00d4ff;'>
                <b>{a.get('title','')[:100]}</b> {lnk}<br>
                <small style='color:#5a6a7a;'>{a.get('domain','')} · {a.get('query_label','')} · {a.get('snippet','')[:120]}</small>
                </div>""", unsafe_allow_html=True)
            if art_other:
                st.markdown(f"**Autres articles ({len(art_other)})**")
                for a in art_other[:30]:
                    url = a.get("url","")
                    lnk = f"<a href='{url}' target='_blank' style='color:#5a6a7a;'>→ lire</a>" if url else ""
                    st.markdown(f"""<div class='result-row'>
                    <span style='color:#5a6a7a;'>{a.get('title','')[:100]}</span> {lnk}
                    </div>""", unsafe_allow_html=True)

        # ── HUMAN VALIDATION ──────────────────────────────────────
        st.markdown("---")
        st.markdown("<div class='section-title'>✍️ VALIDATION HUMAINE</div>", unsafe_allow_html=True)
        st.markdown("<div class='info-box'>Après examen des articles, renseignez votre décision. Le rapport PDF et l'historique Excel seront horodatés avec votre validation.</div>", unsafe_allow_html=True)

        vh1, vh2 = st.columns([1,2])
        with vh1:
            analyst_name = st.text_input("👤 Nom de l'analyste", placeholder="ex: Marie Dupont", key="analyst_name")
            human_decision_raw = st.radio(
                "Décision après analyse",
                ["En attente", "✅ RAS — Rien à signaler", "⚠️ Informations négatives confirmées"],
                key="human_decision")
        with vh2:
            human_comment = st.text_area("Commentaire", placeholder="ex: RAS — homonyme identifié.", height=110, key="human_comment")

        decision_map = {"En attente":None, "✅ RAS — Rien à signaler":"RAS",
                        "⚠️ Informations négatives confirmées":"RISQUE_CONFIRME"}
        h_decision = decision_map.get(human_decision_raw, None)

        # PDF + Excel buttons
        st.markdown("")
        pdf_c1, pdf_c2, pdf_c3 = st.columns(3)

        with pdf_c1:
            if st.button("⬇ GÉNÉRER RAPPORT PDF OSINT", key="gen_pdf_main"):
                with st.spinner("Génération du PDF…"):
                    try:
                        pdf_bytes = generate_osint_pdf(
                            entity_d, analysis, os_result,
                            iban_data if iban_data.get("raw") else None,
                            bank_data if bank_data else None,
                            human_decision=h_decision,
                            human_comment=human_comment,
                            analyst_name=analyst_name
                        )
                        suffix = {"RAS":"RAS","RISQUE_CONFIRME":"RISQUE"}.get(h_decision,"ATTENTE")
                        fname  = f"FinShield_OSINT_{entity_d.replace(' ','_')}_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                        st.download_button("📥 Télécharger PDF", data=pdf_bytes,
                                           file_name=fname, mime="application/pdf",
                                           key="dl_pdf_main")
                        st.markdown(f"<div class='ok-box'>✅ Rapport PDF généré.</div>", unsafe_allow_html=True)
                    except Exception as e:
                        import traceback
                        st.error(f"Erreur PDF : {e}")
                        st.code(traceback.format_exc())

        with pdf_c2:
            if st.button("📊 METTRE À JOUR EXCEL (avec validation)", key="update_excel"):
                try:
                    xl_path = append_to_excel_history(
                        entity_d, entity_type, analysis, os_result,
                        iban_data if iban_data.get("raw") else None,
                        analyst_name=analyst_name, human_decision=h_decision)
                    with open(xl_path, "rb") as f:
                        st.download_button("📥 Télécharger Excel", data=f.read(),
                                           file_name="finshield_history.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="dl_excel_main")
                    st.markdown("<div class='ok-box'>✅ Historique Excel mis à jour.</div>", unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Erreur Excel : {e}")

        with pdf_c3:
            if os.path.exists(EXCEL_HISTORY_PATH):
                with open(EXCEL_HISTORY_PATH, "rb") as f:
                    st.download_button("📥 Télécharger l'historique Excel complet",
                                       data=f.read(),
                                       file_name="finshield_history.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_excel_full")


# ─────────────────────────────────────────────────────────────────
# § 10  TAB 3 — BANK SEARCH
# ─────────────────────────────────────────────────────────────────
with tab3:
    st.markdown("## Recherche de Banque")
    st.markdown("<div class='info-box'>Recherche par code CIB, nom, BIC ou ville.</div>", unsafe_allow_html=True)
    sc1, sc2 = st.columns([3,1])
    with sc1:
        bq = st.text_input("Code CIB, nom, BIC…", placeholder="30004 · BNP · BNPAFRPP", key="bsearch")
    with sc2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        do_search = st.button("▶ RECHERCHER", key="btn_bsearch")
    if do_search and bq:
        hits = db_get_banks(bq)
        st.markdown(f"<div class='info-box'>{len(hits)} résultat(s)</div>", unsafe_allow_html=True)
        for b in hits[:20]:
            bic_str = f" · BIC : <code>{b['bic']}</code>" if b.get("bic") else ""
            st.markdown(f"""<div class='metric-card'>
            <span style='font-family:IBM Plex Mono,monospace;color:#00d4ff;'>{b['code']}</span>{bic_str}
            <div style='font-size:0.95rem;color:#c8d6e5;margin-top:6px;'><b>{b['name']}</b></div>
            <div style='font-size:0.8rem;color:#5a6a7a;'>{b.get('city','')} {b.get('postal_code','')} · {b.get('country','')}</div>
            </div>""", unsafe_allow_html=True)
    with st.expander("📖 Table complète"):
        flt = st.text_input("Filtrer", key="full_bflt").lower()
        rows = db_get_banks(flt)
        if rows:
            df = pd.DataFrame(rows)[["code","name","bic","type","city","country"]]
            df.columns = ["Code CIB","Banque","BIC","Type","Ville","Pays"]
            st.dataframe(df, use_container_width=True, height=450)


# ─────────────────────────────────────────────────────────────────
# § 11  TAB 4 — DATABASE MANAGEMENT
# ─────────────────────────────────────────────────────────────────
with tab4:
    st.markdown("## Gestion de la Base de Données")
    db_tab1, db_tab2, db_tab3 = st.tabs(["🏦 Banques", "🌍 Pays IBAN", "📤 Import / Export"])

    with db_tab1:
        st.markdown("### Ajouter / Modifier une banque")
        with st.form("form_bank"):
            fc1,fc2,fc3 = st.columns(3)
            with fc1:
                f_code = st.text_input("Code CIB *"); f_name = st.text_input("Nom *")
                f_bic  = st.text_input("BIC / SWIFT")
            with fc2:
                f_addr = st.text_input("Adresse"); f_city = st.text_input("Ville")
                f_cp   = st.text_input("Code postal")
            with fc3:
                f_country = st.selectbox("Pays", ["FR","BE","LU","MC","CH","DE","GB","ES","IT","NL","PT","Other"])
                f_type    = st.selectbox("Type", ["Etablissement de crédit","Banque centrale","Autre institution"])
                f_notes   = st.text_area("Notes", height=68)
            if st.form_submit_button("💾 ENREGISTRER"):
                if f_code and f_name:
                    db_upsert_bank(f_code.strip(),f_name.strip(),f_addr.strip(),
                                   f_city.strip(),f_cp.strip(),f_country,
                                   f_bic.strip().upper(),f_type,f_notes.strip())
                    st.success(f"✅ Banque '{f_name}' (code {f_code}) enregistrée.")
                    st.rerun()
                else: st.error("Code CIB et nom obligatoires.")

        filter_banks = st.text_input("Filtrer la liste", key="db_bank_flt")
        for b in db_get_banks(filter_banks)[:100]:
            col_b1, col_b2 = st.columns([5,1])
            with col_b1:
                st.markdown(f"""<div class='result-row'>
                <b style='color:#00d4ff;'>{b['code']}</b>{"  ·  "+b['bic'] if b.get('bic') else ""} · {b['name']}
                <span style='color:#5a6a7a;font-size:0.8rem;'> · {b.get('city','')} · {b.get('type','')}</span>
                </div>""", unsafe_allow_html=True)
            with col_b2:
                if st.button(f"🗑", key=f"del_{b['code']}"):
                    db_delete_bank(b["code"]); st.rerun()

    with db_tab2:
        st.markdown("### Ajouter / Modifier un pays IBAN")
        with st.form("form_iban_country"):
            ic1,ic2,ic3 = st.columns(3)
            with ic1:
                ic_code=st.text_input("Code ISO *",max_chars=2); ic_name=st.text_input("Nom *")
                ic_length=st.number_input("Longueur *",min_value=15,max_value=34,value=27)
            with ic2:
                ic_bban=st.text_input("Format BBAN"); ic_struct=st.text_input("Structure")
            with ic3:
                ic_example=st.text_input("Exemple IBAN"); ic_notes=st.text_area("Notes",height=68)
            if st.form_submit_button("💾 ENREGISTRER"):
                if ic_code and ic_name:
                    db_upsert_iban_country(ic_code.upper(),ic_name,ic_length,ic_struct,ic_example,ic_bban,ic_notes)
                    st.success(f"✅ {ic_code.upper()} — {ic_name} enregistré."); st.rerun()
                else: st.error("Code et nom obligatoires.")
        df_ic = pd.DataFrame(db_get_all_iban_countries())[["code","name","length","bban_format","example"]]
        df_ic.columns = ["Code","Pays","Longueur","Format BBAN","Exemple"]
        st.dataframe(df_ic, use_container_width=True, height=400)

    with db_tab3:
        exp1,exp2 = st.columns(2)
        with exp1:
            st.markdown("#### 📤 Export banques (CSV)")
            banks_all = db_get_banks()
            if banks_all:
                buf2 = io.StringIO()
                csv.DictWriter(buf2, fieldnames=banks_all[0].keys()).writeheader()
                csv.DictWriter(buf2, fieldnames=banks_all[0].keys()).writerows(banks_all)
                st.download_button("⬇ banks.csv", data=buf2.getvalue().encode(),
                                   file_name="finshield_banks.csv", mime="text/csv")
        with exp2:
            st.markdown("#### 📥 Import banques (CSV)")
            uploaded = st.file_uploader("CSV (colonnes: code, name, …)", type=["csv"], key="import_banks")
            if uploaded:
                try:
                    df_up = pd.read_csv(uploaded)
                    df_up.columns = [c.lower().strip() for c in df_up.columns]
                    if "code" in df_up.columns and "name" in df_up.columns:
                        st.dataframe(df_up.head(), use_container_width=True)
                        if st.button("✅ Confirmer l'import"):
                            for _, row in df_up.iterrows():
                                db_upsert_bank(str(row.get("code","")).strip(),
                                               str(row.get("name","")).strip(),
                                               str(row.get("address","")).strip(),
                                               str(row.get("city","")).strip(),
                                               str(row.get("postal_code","")).strip(),
                                               str(row.get("country","FR")).strip(),
                                               str(row.get("bic","")).strip().upper(),
                                               str(row.get("type","Etablissement de crédit")).strip(),
                                               str(row.get("notes","")).strip())
                            st.success(f"✅ {len(df_up)} banques importées."); st.rerun()
                    else: st.error("Colonnes 'code' et 'name' requises.")
                except Exception as e: st.error(f"Erreur : {e}")

        st.markdown("---")
        with st.expander("⚠️ Réinitialisation"):
            cr1,cr2,cr3 = st.columns(3)
            with cr1:
                if st.button("🗑 Vider l'historique"):
                    conn=get_db(); conn.execute("DELETE FROM osint_reports"); conn.commit(); conn.close()
                    st.success("Vidé."); st.rerun()
            with cr2:
                if st.button("🗑 Vider la watchlist"):
                    conn=get_db(); conn.execute("DELETE FROM watchlist"); conn.commit(); conn.close()
                    st.success("Vidée."); st.rerun()
            with cr3:
                if st.button("🔄 Réensemencer banques"):
                    conn=get_db(); conn.execute("DELETE FROM banks"); conn.commit(); conn.close()
                    seed_banks(); st.success("Réinitialisé."); st.rerun()


# ─────────────────────────────────────────────────────────────────
# § 12  TAB 5 — HISTORY & WATCHLIST
# ─────────────────────────────────────────────────────────────────
with tab5:
    st.markdown("## Historique & Surveillance")
    ht1, ht2, ht3 = st.tabs(["📋 Rapports", "👁 Watchlist", "📊 Excel Historique"])

    with ht1:
        reports = db_get_reports(100)
        if reports:
            df_rep = pd.DataFrame(reports)[["id","entity","entity_type","score","niveau","recommandation","created_at"]]
            df_rep.columns = ["ID","Entité","Type","Score","Niveau","Recommandation","Date"]
            st.dataframe(df_rep, use_container_width=True, height=300)
            sel = st.selectbox("Voir le détail", ["—"] + [f"#{r['id']} — {r['entity']} ({r['created_at'][:10]})" for r in reports])
            if sel != "—":
                idx = int(sel.split("—")[0].replace("#","").strip()) - 1
                rep = reports[idx]
                st.markdown(f"""<div class='metric-card'>
                <div class='label'>Entité</div><div class='value' style='font-size:1rem;'>{rep['entity']}</div>
                <div class='sub'>Score {rep['score']}/100 · {rep['niveau']} · {rep['recommandation']}</div>
                <div class='sub'>{rep['resume']}</div></div>""", unsafe_allow_html=True)
                if rep.get("full_json"):
                    with st.expander("JSON complet"):
                        try: st.json(json.loads(rep["full_json"]))
                        except: st.text(rep["full_json"])
                if st.button("⬇ Régénérer PDF", key=f"repdf_{rep['id']}"):
                    try:
                        analysis = json.loads(rep.get("full_json","{}"))
                        pdf = generate_osint_pdf(rep["entity"], analysis,
                                                  analysis.get("os_result",{"count":0,"results":[]}))
                        st.download_button("📥 Télécharger",data=pdf,
                            file_name=f"FinShield_{rep['entity'].replace(' ','_')}_{rep['created_at'][:10]}.pdf",
                            mime="application/pdf")
                    except Exception as e: st.error(f"Erreur : {e}")
        else:
            st.markdown("<div class='info-box'>Aucun rapport. Lancez une analyse OSINT.</div>", unsafe_allow_html=True)

    with ht2:
        st.markdown("### Ajouter une entité")
        with st.form("form_watch"):
            wc1,wc2 = st.columns(2)
            with wc1:
                w_entity=st.text_input("Entité *"); w_type=st.selectbox("Type",["Entreprise","Personne","Autre"])
            with wc2:
                w_reason=st.text_area("Motif",height=68); w_risk=st.selectbox("Risque",["FAIBLE","MODERE","ELEVE","CRITIQUE"])
            w_by=st.text_input("Ajouté par")
            if st.form_submit_button("➕ AJOUTER") and w_entity:
                db_add_watchlist(w_entity,w_type,w_reason,w_risk,w_by); st.success("✅ Ajouté."); st.rerun()
        for w in db_get_watchlist():
            wc1,wc2 = st.columns([5,1])
            risk_c = {"FAIBLE":"badge-low","MODERE":"badge-medium","ELEVE":"badge-high","CRITIQUE":"badge-high"}.get(w["risk_level"],"badge-medium")
            with wc1:
                st.markdown(f"""<div class='result-row'>
                <b>{w['entity']}</b> · <span class='{risk_c}'>{w['risk_level']}</span><br>
                <small style='color:#5a6a7a;'>{w.get('entity_type','')} · {w['created_at'][:10]} · {w.get('added_by','')}</small><br>
                <small>{w.get('reason','')}</small></div>""", unsafe_allow_html=True)
            with wc2:
                if st.button("🗑", key=f"del_watch_{w['id']}"):
                    db_delete_watchlist(w["id"]); st.rerun()

    with ht3:
        st.markdown("### Historique Excel")
        st.markdown("""<div class='info-box'>
        Le fichier Excel (<code>finshield_history.xlsx</code>) contient 3 onglets :
        <b>Historique</b> (une ligne par analyse), <b>Alertes</b> (détail de chaque signal),
        <b>Sources</b> (toutes les pages collectées).
        Il est mis à jour automatiquement à chaque analyse.</div>""", unsafe_allow_html=True)
        if os.path.exists(EXCEL_HISTORY_PATH):
            with open(EXCEL_HISTORY_PATH,"rb") as f:
                st.download_button("📥 Télécharger finshield_history.xlsx",
                                   data=f.read(),
                                   file_name="finshield_history.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            try:
                wb_preview = load_workbook(EXCEL_HISTORY_PATH, data_only=True)
                ws = wb_preview["Historique"]
                data = list(ws.values)
                if len(data) > 1:
                    df_prev = pd.DataFrame(data[1:], columns=data[0])
                    st.dataframe(df_prev, use_container_width=True, height=350)
            except Exception as e: st.warning(f"Aperçu non disponible : {e}")
        else:
            st.markdown("<div class='warn-box'>Aucun fichier Excel encore créé. Lancez une analyse OSINT.</div>", unsafe_allow_html=True)
